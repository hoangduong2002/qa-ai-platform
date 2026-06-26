from pathlib import Path

from openpyxl import load_workbook

from app.exporters.function_based_excel_exporter import (
    export_function_based_testcases_to_excel,
)


def _summary_value(ws, metric: str):
    for row in ws.iter_rows(values_only=True):
        if row and row[0] == metric:
            return row[1]
    raise AssertionError(f"Metric not found: {metric}")


def test_excel_export_includes_automation_sheets_and_counts(tmp_path, monkeypatch):
    monkeypatch.chdir(tmp_path)

    testcases = [
        {
            "testcase_id": "TC001",
            "function_id": "FUNC001",
            "scenario_id": "SC001",
            "title": "Login succeeds",
            "priority": "High",
            "preconditions": ["User exists"],
            "test_steps": ["Navigate to login", "Input credentials", "Click Login"],
            "expected_results": ["Dashboard is displayed"],
            "related_requirement_ids": ["REQ001"],
            "execution_type": "AUTOMATION",
            "automation_candidate": True,
            "automation_tool": "Playwright",
            "automation_priority": "High",
            "automation_reason": "Browser UI flow with deterministic assertion.",
            "automation_blockers": [],
            "manual_reason": "",
        },
        {
            "testcase_id": "TC002",
            "function_id": "FUNC001",
            "scenario_id": "SC002",
            "title": "Review visual layout",
            "priority": "Medium",
            "preconditions": ["Dashboard is available"],
            "test_steps": ["Open dashboard", "Review visual layout and color"],
            "expected_results": ["Layout is acceptable"],
            "related_requirement_ids": ["REQ002"],
            "execution_type": "MANUAL",
            "automation_candidate": False,
            "automation_tool": "",
            "automation_priority": "Not Applicable",
            "automation_reason": "",
            "automation_blockers": ["visual", "layout", "color"],
            "manual_reason": "Requires human visual review.",
        },
        {
            "testcase_id": "TC003",
            "function_id": "FUNC001",
            "scenario_id": "SC003",
            "title": "Submit request and complete manual approval verification",
            "priority": "Low",
            "preconditions": ["Request form is available"],
            "test_steps": [
                "Navigate to request form",
                "Input valid request details",
                "Click Submit",
                "Perform manual verification of approval email",
            ],
            "expected_results": [
                "Request confirmation is displayed",
                "Approval email is manually verified",
            ],
            "related_requirement_ids": ["REQ003"],
            "execution_type": "HYBRID",
            "automation_candidate": True,
            "automation_tool": "Playwright",
            "automation_priority": "Low",
            "automation_reason": "UI submission can be automated.",
            "automation_blockers": ["manual verification", "email inbox"],
            "manual_reason": "Final approval email must be manually verified.",
        },
    ]

    export_path = export_function_based_testcases_to_excel(
        ticket_id="TEST-AUTO-EXPORT",
        testcases=testcases,
        coverage_review={},
        final_coverage_review={},
        approved_structure={
            "main_functions": [
                {
                    "function_id": "FUNC001",
                    "name": "Login and Requests",
                    "related_requirement_ids": ["REQ001", "REQ002", "REQ003"],
                }
            ]
        },
        analysis={
            "requirement_items": [
                {"requirement_id": "REQ001", "type": "Functional", "description": "Login"},
                {"requirement_id": "REQ002", "type": "UX", "description": "Layout"},
                {"requirement_id": "REQ003", "type": "Workflow", "description": "Approval"},
            ]
        },
    )

    workbook_path = Path(export_path)
    assert workbook_path.exists()

    wb = load_workbook(workbook_path)
    assert {
        "All Test Cases",
        "Automation Candidates",
        "Manual Test Cases",
        "Automation Summary",
    }.issubset(set(wb.sheetnames))

    assert wb["All Test Cases"].max_row == 4
    assert wb["Automation Candidates"].max_row == 3
    assert wb["Manual Test Cases"].max_row == 3

    summary = wb["Automation Summary"]
    assert _summary_value(summary, "Total test cases") == "3"
    assert _summary_value(summary, "Automation candidates") == "2"
    assert _summary_value(summary, "Manual test cases") == "1"
    assert _summary_value(summary, "Hybrid test cases") == "1"
    assert _summary_value(summary, "High priority automation count") == "1"
