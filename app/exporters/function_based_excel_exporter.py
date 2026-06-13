import json
import re
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.utils.test_structure_store import load_approved_test_case_structure


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
SECTION_FILL = PatternFill("solid", fgColor="D9EAF7")
BOLD_FONT = Font(bold=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)


def _safe_sheet_name(name: str) -> str:
    invalid_chars = ["\\", "/", "*", "[", "]", ":", "?"]

    clean_name = name or "Sheet"

    for char in invalid_chars:
        clean_name = clean_name.replace(char, "_")

    clean_name = clean_name.strip()

    if not clean_name:
        clean_name = "Sheet"

    return clean_name[:31]


def _to_text(value: Any) -> str:
    if value is None:
        return ""

    if isinstance(value, str):
        return value

    if isinstance(value, list):
        return "\n".join(_to_text(item) for item in value)

    if isinstance(value, dict):
        return json.dumps(value, ensure_ascii=False, indent=2)

    return str(value)


def _ensure_exports_dir(ticket_id: str) -> Path:
    exports_dir = Path("requirements") / ticket_id / "exports"
    exports_dir.mkdir(parents=True, exist_ok=True)
    return exports_dir


def _apply_table_style(ws, freeze_row: int = 1) -> None:
    ws.freeze_panes = f"A{freeze_row + 1}"
    ws.auto_filter.ref = ws.dimensions

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )
            cell.border = THIN_BORDER

    for cell in ws[freeze_row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )


def _auto_width(ws, max_width: int = 60) -> None:
    for column_cells in ws.columns:
        column_letter = get_column_letter(column_cells[0].column)
        max_length = 0

        for cell in column_cells:
            value = cell.value

            if value is None:
                continue

            lines = str(value).splitlines() or [str(value)]

            for line in lines:
                max_length = max(max_length, len(line))

        ws.column_dimensions[column_letter].width = min(
            max(max_length + 2, 12),
            max_width,
        )


def _append_kv(ws, key: str, value: Any) -> None:
    row = ws.max_row + 1
    ws.cell(row=row, column=1, value=key)
    ws.cell(row=row, column=2, value=_to_text(value))
    ws.cell(row=row, column=1).font = BOLD_FONT
    ws.cell(row=row, column=1).fill = SECTION_FILL


def _get_functions(approved_structure: dict) -> list[dict]:
    if not isinstance(approved_structure, dict):
        return []

    functions = (
        approved_structure.get("main_functions")
        or approved_structure.get("functions")
        or approved_structure.get("test_functions")
        or []
    )

    if not isinstance(functions, list):
        return []

    return functions


def _get_function_id(function_item: dict) -> str:
    if not isinstance(function_item, dict):
        return ""

    return (
        function_item.get("function_id")
        or function_item.get("main_function_id")
        or function_item.get("id")
        or ""
    )


def _get_function_name(function_item: dict) -> str:
    if not isinstance(function_item, dict):
        return ""

    return (
        function_item.get("name")
        or function_item.get("title")
        or function_item.get("function_name")
        or ""
    )


def _get_sub_functions(function_item: dict) -> list[dict]:
    if not isinstance(function_item, dict):
        return []

    sub_functions = (
        function_item.get("sub_functions")
        or function_item.get("subfunctions")
        or function_item.get("children")
        or []
    )

    if not isinstance(sub_functions, list):
        return []

    return sub_functions


def _get_sub_function_id(sub_function_item: dict) -> str:
    if not isinstance(sub_function_item, dict):
        return ""

    return (
        sub_function_item.get("sub_function_id")
        or sub_function_item.get("subfunction_id")
        or sub_function_item.get("id")
        or ""
    )


def _get_sub_function_name(sub_function_item: dict) -> str:
    if not isinstance(sub_function_item, dict):
        return ""

    return (
        sub_function_item.get("name")
        or sub_function_item.get("title")
        or sub_function_item.get("sub_function_name")
        or ""
    )


def _get_test_areas(sub_function_item: dict) -> list[dict]:
    if not isinstance(sub_function_item, dict):
        return []

    test_areas = (
        sub_function_item.get("test_areas")
        or sub_function_item.get("detailed_test_areas")
        or sub_function_item.get("areas")
        or []
    )

    if not isinstance(test_areas, list):
        return []

    return test_areas


def _get_test_area_id(test_area_item: dict) -> str:
    if not isinstance(test_area_item, dict):
        return ""

    return (
        test_area_item.get("test_area_id")
        or test_area_item.get("area_id")
        or test_area_item.get("id")
        or ""
    )


def _get_test_area_name(test_area_item: dict) -> str:
    if not isinstance(test_area_item, dict):
        return ""

    return (
        test_area_item.get("name")
        or test_area_item.get("title")
        or test_area_item.get("test_area_name")
        or ""
    )


def _get_related_requirement_ids_from_structure(item: dict) -> list[str]:
    if not isinstance(item, dict):
        return []

    value = (
        item.get("related_requirement_ids")
        or item.get("requirement_ids")
        or item.get("related_requirements")
        or []
    )

    if isinstance(value, str):
        return [
            requirement_id.strip()
            for requirement_id in value.split(",")
            if requirement_id.strip()
        ]

    if isinstance(value, list):
        return [
            str(requirement_id).strip()
            for requirement_id in value
            if str(requirement_id).strip()
        ]

    return []


def _get_item_function_id(item: dict) -> str:
    if not isinstance(item, dict):
        return ""

    return (
        item.get("function_id")
        or item.get("main_function_id")
        or ""
    )


def _group_testcases_by_function(
    testcases: list,
    approved_structure: dict,
) -> dict[str, dict]:
    groups = {}

    for function_item in _get_functions(approved_structure):
        function_id = _get_function_id(function_item)

        if not function_id:
            continue

        groups[function_id] = {
            "function": function_item,
            "testcases": [],
        }

    for testcase in testcases:
        if not isinstance(testcase, dict):
            continue

        function_id = _get_item_function_id(testcase)

        if not function_id:
            function_id = "UNMAPPED"

        if function_id not in groups:
            groups[function_id] = {
                "function": {
                    "function_id": function_id,
                    "name": (
                        "Unmapped Test Cases"
                        if function_id == "UNMAPPED"
                        else function_id
                    ),
                },
                "testcases": [],
            }

        groups[function_id]["testcases"].append(testcase)

    return groups


def _write_testcase_header(ws) -> None:
    headers = [
        "Test Case ID",
        "Scenario ID",
        "Function ID",
        "Sub Function ID",
        "Test Area ID",
        "Title",
        "Type",
        "Technique",
        "Priority",
        "Preconditions",
        "Test Steps",
        "Expected Results",
        "Traceability",
        "Change Status",
        "Previous Test Case ID",
        "Related Change IDs",
        "Source Snapshot Version",
    ]

    ws.append(headers)


def _write_testcase_row(ws, testcase: dict) -> None:
    ws.append(
        [
            testcase.get("testcase_id", ""),
            testcase.get("scenario_id", ""),
            testcase.get("function_id", ""),
            testcase.get("sub_function_id", ""),
            testcase.get("test_area_id", ""),
            testcase.get("title", ""),
            testcase.get("type", ""),
            testcase.get("technique", ""),
            testcase.get("priority", ""),
            _to_text(testcase.get("preconditions", [])),
            _to_text(testcase.get("test_steps", [])),
            _to_text(testcase.get("expected_results", [])),
            testcase.get("traceability", ""),
            testcase.get("change_status", ""),
            testcase.get("previous_testcase_id", ""),
            _to_text(testcase.get("related_change_ids", [])),
            testcase.get("source_snapshot_version", ""),
        ]
    )


def _create_summary_sheet(
    wb: Workbook,
    ticket_id: str,
    approved_structure: dict,
    testcases: list,
    coverage_review: dict,
    final_coverage_review: dict,
) -> None:
    ws = wb.active
    ws.title = "Summary"

    ws.append(["Field", "Value"])

    _append_kv(ws, "Requirement ID", ticket_id)
    _append_kv(ws, "Generation Mode", "Function-Based Structured Generation")
    _append_kv(ws, "Main Functions", len(_get_functions(approved_structure)))
    _append_kv(ws, "Total Test Cases", len(testcases))
    _append_kv(ws, "Coverage Score", coverage_review.get("coverage_score", ""))
    _append_kv(ws, "Coverage Approved By AI", coverage_review.get("approved_by_ai", ""))
    _append_kv(
        ws,
        "Final Coverage Score",
        final_coverage_review.get("final_coverage_score")
        or final_coverage_review.get("coverage_score")
        or "",
    )
    _append_kv(
        ws,
        "Ready For Execution",
        final_coverage_review.get("ready_for_execution", ""),
    )
    _append_kv(
        ws,
        "Final Approved By AI",
        final_coverage_review.get("approved_by_ai", ""),
    )

    _apply_table_style(ws)
    _auto_width(ws)


def _create_test_case_structure_sheet(
    wb: Workbook,
    approved_structure: dict,
) -> None:
    ws = wb.create_sheet("Test Case Structure")

    ws.append(
        [
            "Function ID",
            "Function Name",
            "Sub Function ID",
            "Sub Function Name",
            "Test Area ID",
            "Test Area Name",
            "Description",
            "Related Requirement IDs",
            "Traceability",
            "Priority",
            "Included",
            "Notes",
        ]
    )

    functions = _get_functions(approved_structure)

    for function_item in functions:
        function_id = _get_function_id(function_item)
        function_name = _get_function_name(function_item)

        function_requirement_ids = _get_related_requirement_ids_from_structure(
            function_item
        )

        sub_functions = _get_sub_functions(function_item)

        if not sub_functions:
            ws.append(
                [
                    function_id,
                    function_name,
                    "",
                    "",
                    "",
                    "",
                    _to_text(
                        function_item.get("description")
                        or function_item.get("summary")
                        or ""
                    ),
                    _to_text(function_requirement_ids),
                    ", ".join(function_requirement_ids),
                    function_item.get("priority", ""),
                    function_item.get("included", True),
                    _to_text(function_item.get("notes", "")),
                ]
            )
            continue

        for sub_function_item in sub_functions:
            sub_function_id = _get_sub_function_id(sub_function_item)
            sub_function_name = _get_sub_function_name(sub_function_item)

            sub_function_requirement_ids = (
                _get_related_requirement_ids_from_structure(sub_function_item)
                or function_requirement_ids
            )

            test_areas = _get_test_areas(sub_function_item)

            if not test_areas:
                ws.append(
                    [
                        function_id,
                        function_name,
                        sub_function_id,
                        sub_function_name,
                        "",
                        "",
                        _to_text(
                            sub_function_item.get("description")
                            or sub_function_item.get("summary")
                            or ""
                        ),
                        _to_text(sub_function_requirement_ids),
                        ", ".join(sub_function_requirement_ids),
                        sub_function_item.get("priority", ""),
                        sub_function_item.get("included", True),
                        _to_text(sub_function_item.get("notes", "")),
                    ]
                )
                continue

            for test_area_item in test_areas:
                test_area_id = _get_test_area_id(test_area_item)
                test_area_name = _get_test_area_name(test_area_item)

                test_area_requirement_ids = (
                    _get_related_requirement_ids_from_structure(test_area_item)
                    or sub_function_requirement_ids
                )

                ws.append(
                    [
                        function_id,
                        function_name,
                        sub_function_id,
                        sub_function_name,
                        test_area_id,
                        test_area_name,
                        _to_text(
                            test_area_item.get("description")
                            or test_area_item.get("summary")
                            or test_area_item.get("test_objective")
                            or ""
                        ),
                        _to_text(test_area_requirement_ids),
                        ", ".join(test_area_requirement_ids),
                        test_area_item.get("priority", ""),
                        test_area_item.get("included", True),
                        _to_text(test_area_item.get("notes", "")),
                    ]
                )

    _apply_table_style(ws)
    _auto_width(ws)


def _create_master_testcases_sheet(wb: Workbook, testcases: list) -> None:
    ws = wb.create_sheet("Master Test Cases")
    _write_testcase_header(ws)

    for testcase in testcases:
        _write_testcase_row(ws, testcase)

    _apply_table_style(ws)
    _auto_width(ws)


def _create_function_testcase_sheets(
    wb: Workbook,
    grouped_testcases: dict[str, dict],
) -> None:
    for function_id, group in grouped_testcases.items():
        function_item = group["function"]
        function_name = _get_function_name(function_item)

        sheet_name = _safe_sheet_name(function_id)

        ws = wb.create_sheet(sheet_name)

        ws.append(["Function ID", function_id])
        ws.append(["Function Name", function_name])
        ws.append([])
        _write_testcase_header(ws)

        for testcase in group["testcases"]:
            _write_testcase_row(ws, testcase)

        _apply_table_style(ws, freeze_row=4)
        _auto_width(ws)


def _create_coverage_review_sheet(
    wb: Workbook,
    coverage_review: dict,
) -> None:
    ws = wb.create_sheet("Coverage Review")

    ws.append(
        [
            "Function ID",
            "Function Name",
            "Coverage Score",
            "Approved By AI",
            "Scenario Count",
            "Test Case Count",
            "Summary",
            "Missing Scenarios",
            "Weak Test Cases",
            "Missing Test Cases",
            "Traceability Issues",
            "Recommendations",
        ]
    )

    function_reviews = coverage_review.get("function_reviews", [])

    if function_reviews:
        for review in function_reviews:
            ws.append(
                [
                    review.get("function_id", ""),
                    review.get("function_name", ""),
                    review.get("coverage_score", ""),
                    review.get("approved_by_ai", ""),
                    review.get("scenario_count", ""),
                    review.get("testcase_count", ""),
                    review.get("summary", ""),
                    _to_text(review.get("missing_scenarios", [])),
                    _to_text(review.get("weak_testcases", [])),
                    _to_text(review.get("missing_testcases", [])),
                    _to_text(review.get("traceability_issues", [])),
                    _to_text(review.get("recommendations", [])),
                ]
            )
    else:
        ws.append(
            [
                "",
                "",
                coverage_review.get("coverage_score", ""),
                coverage_review.get("approved_by_ai", ""),
                coverage_review.get("scenario_count", ""),
                coverage_review.get("testcase_count", ""),
                coverage_review.get("summary", ""),
                _to_text(coverage_review.get("missing_scenarios", [])),
                _to_text(coverage_review.get("weak_testcases", [])),
                _to_text(coverage_review.get("missing_testcases", [])),
                _to_text(coverage_review.get("traceability_issues", [])),
                _to_text(coverage_review.get("recommendations", [])),
            ]
        )

    _apply_table_style(ws)
    _auto_width(ws)


def _create_final_review_sheet(
    wb: Workbook,
    final_coverage_review: dict,
) -> None:
    ws = wb.create_sheet("Final Review")

    ws.append(
        [
            "Function ID",
            "Function Name",
            "Final Coverage Score",
            "Approved By AI",
            "Ready For Execution",
            "Scenario Count",
            "Test Case Count",
            "Summary",
            "Remaining Gaps",
            "Traceability Issues",
            "Execution Readiness Issues",
            "Final Recommendations",
        ]
    )

    function_reviews = final_coverage_review.get("function_reviews", [])

    if function_reviews:
        for review in function_reviews:
            ws.append(
                [
                    review.get("function_id", ""),
                    review.get("function_name", ""),
                    review.get("final_coverage_score", ""),
                    review.get("approved_by_ai", ""),
                    review.get("ready_for_execution", ""),
                    review.get("scenario_count", ""),
                    review.get("testcase_count", ""),
                    review.get("summary", ""),
                    _to_text(review.get("remaining_gaps", [])),
                    _to_text(review.get("traceability_issues", [])),
                    _to_text(review.get("execution_readiness_issues", [])),
                    _to_text(review.get("final_recommendations", [])),
                ]
            )
    else:
        ws.append(
            [
                "",
                "",
                final_coverage_review.get("final_coverage_score")
                or final_coverage_review.get("coverage_score", ""),
                final_coverage_review.get("approved_by_ai", ""),
                final_coverage_review.get("ready_for_execution", ""),
                final_coverage_review.get("scenario_count", ""),
                final_coverage_review.get("testcase_count", ""),
                final_coverage_review.get("summary", ""),
                _to_text(final_coverage_review.get("remaining_gaps", [])),
                _to_text(final_coverage_review.get("traceability_issues", [])),
                _to_text(final_coverage_review.get("execution_readiness_issues", [])),
                _to_text(final_coverage_review.get("final_recommendations", [])),
            ]
        )

    _apply_table_style(ws)
    _auto_width(ws)


def _build_requirement_coverage_index(
    testcases: list,
) -> dict:
    coverage = {}

    for testcase in testcases:
        if not isinstance(testcase, dict):
            continue

        requirement_ids = testcase.get("related_requirement_ids", [])

        if isinstance(requirement_ids, str):
            requirement_ids = [
                item.strip()
                for item in requirement_ids.split(",")
                if item.strip()
            ]

        if not isinstance(requirement_ids, list):
            continue

        for requirement_id in requirement_ids:
            if not requirement_id:
                continue

            requirement_id = str(requirement_id).strip()

            if not requirement_id:
                continue

            if requirement_id not in coverage:
                coverage[requirement_id] = {
                    "scenario_ids": set(),
                    "testcase_ids": set(),
                    "function_ids": set(),
                    "testcase_titles": [],
                }

            coverage_item = coverage[requirement_id]

            scenario_id = testcase.get("scenario_id", "")
            testcase_id = testcase.get("testcase_id", "")
            function_id = testcase.get("function_id", "")
            title = testcase.get("title", "")

            if scenario_id:
                coverage_item["scenario_ids"].add(str(scenario_id))

            if testcase_id:
                coverage_item["testcase_ids"].add(str(testcase_id))

            if function_id:
                coverage_item["function_ids"].add(str(function_id))

            if title:
                coverage_item["testcase_titles"].append(str(title))

    return coverage


def _create_traceability_matrix_sheet(
    wb: Workbook,
    testcases: list,
) -> None:
    ws = wb.create_sheet("Traceability Matrix")

    ws.append(
        [
            "Requirement ID",
            "Function ID",
            "Scenario ID",
            "Test Case ID",
            "Test Case Title",
            "Priority",
            "Type",
        ]
    )

    for testcase in testcases:
        requirement_ids = testcase.get("related_requirement_ids", [])

        if isinstance(requirement_ids, str):
            requirement_ids = [
                item.strip()
                for item in requirement_ids.split(",")
                if item.strip()
            ]

        if not isinstance(requirement_ids, list):
            requirement_ids = []

        for requirement_id in requirement_ids:
            ws.append(
                [
                    requirement_id,
                    testcase.get("function_id", ""),
                    testcase.get("scenario_id", ""),
                    testcase.get("testcase_id", ""),
                    testcase.get("title", ""),
                    testcase.get("priority", ""),
                    testcase.get("type", ""),
                ]
            )

    _apply_table_style(ws)
    _auto_width(ws)
    
    
def _extract_requirement_items(analysis: dict) -> list:
    if not isinstance(analysis, dict):
        return []

    items = (
        analysis.get("requirement_items")
        or analysis.get("requirements")
        or []
    )

    if isinstance(items, list):
        return items

    return []


def _create_requirements_sheet(
    wb: Workbook,
    analysis: dict,
) -> None:
    ws = wb.create_sheet("Requirements")

    ws.append(
        [
            "Requirement ID",
            "Type",
            "Description",
            "Priority",
            "Source",
            "Status",
        ]
    )

    requirement_items = _extract_requirement_items(analysis)

    for item in requirement_items:
        if not isinstance(item, dict):
            continue

        ws.append(
            [
                item.get("requirement_id", ""),
                item.get("type", ""),
                item.get("description", ""),
                item.get("priority", ""),
                item.get("source", ""),
                item.get("status", ""),
            ]
        )

    _apply_table_style(ws)
    _auto_width(ws)
    
    
def _extract_clarification_questions(clarifications: dict) -> list:
    if not isinstance(clarifications, dict):
        return []

    questions = (
        clarifications.get("clarification_questions")
        or clarifications.get("questions")
        or []
    )

    if isinstance(questions, list):
        return questions

    return []


def _extract_answered_clarifications(clarification_answers: dict) -> list:
    if not isinstance(clarification_answers, dict):
        return []

    answered = (
        clarification_answers.get("answered_clarifications")
        or clarification_answers.get("answers")
        or []
    )

    if isinstance(answered, list):
        return answered

    return []


def _create_clarifications_sheet(
    wb: Workbook,
    clarifications: dict,
    clarification_answers: dict,
) -> None:
    ws = wb.create_sheet("Clarifications")

    ws.append(
        [
            "Question ID",
            "Category",
            "Priority",
            "Impact Area",
            "Blocking",
            "Question",
            "Answer",
            "Status",
            "Related Requirement IDs",
            "Reason",
        ]
    )

    questions = _extract_clarification_questions(clarifications)
    answered = _extract_answered_clarifications(clarification_answers)

    answer_by_id = {}

    for item in answered:
        if not isinstance(item, dict):
            continue

        question_id = item.get("question_id", "")

        if question_id:
            answer_by_id[question_id] = item

    for item in questions:
        if not isinstance(item, dict):
            continue

        question_id = item.get("question_id", "")
        answer = answer_by_id.get(question_id, {})

        ws.append(
            [
                question_id,
                item.get("category", ""),
                item.get("priority", item.get("impact", "")),
                item.get("impact_area", ""),
                item.get("blocking", ""),
                item.get("question", ""),
                answer.get("final_answer") or answer.get("answer", ""),
                "Answered" if answer else "Open",
                _to_text(item.get("related_requirement_ids", [])),
                item.get("reason", ""),
            ]
        )

    # Also include answered clarifications that no longer exist in current open questions.
    existing_question_ids = {
        item.get("question_id", "")
        for item in questions
        if isinstance(item, dict)
    }

    for item in answered:
        if not isinstance(item, dict):
            continue

        question_id = item.get("question_id", "")

        if question_id in existing_question_ids:
            continue

        ws.append(
            [
                question_id,
                item.get("category", ""),
                item.get("priority", ""),
                item.get("impact_area", ""),
                item.get("blocking", ""),
                item.get("question", ""),
                item.get("final_answer") or item.get("answer", ""),
                "Answered",
                _to_text(item.get("related_requirement_ids", [])),
                item.get("reason", ""),
            ]
        )

    _apply_table_style(ws)
    _auto_width(ws)
    
    
def _create_requirement_summary_sheet(
    wb: Workbook,
    requirement_summary: dict,
) -> None:
    ws = wb.create_sheet("Requirement Summary")

    ws.append(["Section", "Content"])

    if not isinstance(requirement_summary, dict):
        requirement_summary = {}

    for key in [
        "executive_summary",
        "functional_summary",
        "confirmed_business_rules",
        "validation_rules",
        "open_questions",
        "assumptions",
        "risks",
    ]:
        ws.append(
            [
                key,
                _to_text(requirement_summary.get(key, "")),
            ]
        )

    _apply_table_style(ws)
    _auto_width(ws)


def _create_requirement_matrix_sheet(
    wb: Workbook,
    analysis: dict,
    testcases: list,
) -> None:
    ws = wb.create_sheet("Requirement Matrix")

    ws.append(
        [
            "Requirement ID",
            "Type",
            "Description",
            "Function IDs",
            "Scenario IDs",
            "Test Case IDs",
            "Test Case Count",
            "Coverage Status",
        ]
    )

    requirement_items = _extract_requirement_items(analysis)
    coverage_index = _build_requirement_coverage_index(testcases)

    requirement_ids_from_analysis = set()

    for item in requirement_items:
        if not isinstance(item, dict):
            continue

        requirement_id = (
            item.get("requirement_id")
            or item.get("id")
            or ""
        )

        requirement_id = str(requirement_id).strip()

        if not requirement_id:
            continue

        requirement_ids_from_analysis.add(requirement_id)

        coverage = coverage_index.get(requirement_id, {})

        testcase_ids = sorted(coverage.get("testcase_ids", []))
        scenario_ids = sorted(coverage.get("scenario_ids", []))
        function_ids = sorted(coverage.get("function_ids", []))

        ws.append(
            [
                requirement_id,
                item.get("type", ""),
                item.get("description", ""),
                ", ".join(function_ids),
                ", ".join(scenario_ids),
                ", ".join(testcase_ids),
                len(testcase_ids),
                "Covered" if testcase_ids else "Not Covered",
            ]
        )

    # Include requirement IDs found in test cases but missing from analysis.
    for requirement_id in sorted(
        set(coverage_index.keys()) - requirement_ids_from_analysis
    ):
        coverage = coverage_index.get(requirement_id, {})

        testcase_ids = sorted(coverage.get("testcase_ids", []))
        scenario_ids = sorted(coverage.get("scenario_ids", []))
        function_ids = sorted(coverage.get("function_ids", []))

        ws.append(
            [
                requirement_id,
                "Unknown",
                (
                    "Requirement ID found in test cases but missing from "
                    "analysis.requirement_items"
                ),
                ", ".join(function_ids),
                ", ".join(scenario_ids),
                ", ".join(testcase_ids),
                len(testcase_ids),
                "Covered - Missing Requirement Definition",
            ]
        )

    _apply_table_style(ws)
    _auto_width(ws)


def _create_incremental_testcases_sheet(wb: Workbook, testcases: list) -> None:
    ws = wb.create_sheet("Test Cases")
    _write_testcase_header(ws)

    for testcase in testcases:
        _write_testcase_row(ws, testcase)

    _apply_table_style(ws)
    _auto_width(ws)


def _create_change_impact_sheet(wb: Workbook, change_impact_report: dict) -> None:
    ws = wb.create_sheet("Change Impact Report")
    ws.append(
        [
            "Change ID",
            "Source Type",
            "Source ID",
            "Change Type",
            "Summary",
            "Recommended Action",
            "Impact Confidence",
        ]
    )

    for change in change_impact_report.get("changes", []) or []:
        if not isinstance(change, dict):
            continue
        ws.append(
            [
                change.get("change_id", ""),
                change.get("source_type", ""),
                change.get("source_id", ""),
                change.get("change_type", ""),
                change.get("summary", ""),
                change.get("recommended_action", ""),
                change.get("impact_confidence", ""),
            ]
        )

    _apply_table_style(ws)
    _auto_width(ws)


def _create_regeneration_plan_sheet(wb: Workbook, regeneration_plan: dict) -> None:
    ws = wb.create_sheet("Regeneration Plan")
    ws.append(["Field", "Value"])
    _append_kv(ws, "Impacted Requirements", regeneration_plan.get("impacted_requirement_ids", []))
    _append_kv(ws, "Impacted Scenarios", regeneration_plan.get("impacted_scenario_ids", []))
    _append_kv(ws, "Impacted Test Cases", regeneration_plan.get("impacted_testcase_ids", []))
    _append_kv(
        ws,
        "Can Partial Regenerate",
        regeneration_plan.get("can_partial_regenerate", ""),
    )
    _append_kv(ws, "Confidence", regeneration_plan.get("impact_confidence", ""))
    _append_kv(
        ws,
        "Reason If Full Regenerate Required",
        regeneration_plan.get("reason_if_full_regenerate_required", ""),
    )
    _append_kv(ws, "Source Snapshot Version", regeneration_plan.get("source_snapshot_version", ""))
    _append_kv(ws, "Change Report Version", regeneration_plan.get("change_report_version", ""))

    _apply_table_style(ws)
    _auto_width(ws)


def _status_reason(testcase: dict) -> str:
    status = testcase.get("change_status", "")

    if status == "DeprecatedCandidate":
        return "Linked source was removed or deprecated."

    if status == "Replaced":
        return "Impacted test case retained for history and replaced by regenerated coverage."

    return status


def _create_deprecated_replaced_sheet(wb: Workbook, testcases: list) -> None:
    ws = wb.create_sheet("Deprecated Replaced Test Cases")
    ws.append(
        [
            "Test Case ID",
            "Title",
            "Previous/Current Status",
            "Reason",
            "Related Source Change",
        ]
    )

    for testcase in testcases or []:
        if not isinstance(testcase, dict):
            continue

        status = testcase.get("change_status", "")
        if status not in {"DeprecatedCandidate", "Replaced"}:
            continue

        ws.append(
            [
                testcase.get("testcase_id", ""),
                testcase.get("title", ""),
                status,
                _status_reason(testcase),
                _to_text(testcase.get("related_change_ids", [])),
            ]
        )

    _apply_table_style(ws)
    _auto_width(ws)


def _create_source_traceability_sheet(
    wb: Workbook,
    testcases: list,
    change_impact_report: dict | None = None,
) -> None:
    ws = wb.create_sheet("Source Traceability")
    ws.append(
        [
            "Source Type",
            "Source ID",
            "Requirement ID",
            "Scenario ID",
            "Test Case ID",
            "Source Snapshot Version",
        ]
    )
    change_by_id = {
        change.get("change_id"): change
        for change in (change_impact_report or {}).get("changes", []) or []
        if isinstance(change, dict) and change.get("change_id")
    }

    for testcase in testcases or []:
        if not isinstance(testcase, dict):
            continue

        testcase_id = testcase.get("testcase_id", "")
        scenario_id = testcase.get("related_scenario_id") or testcase.get("scenario_id", "")
        snapshot_version = testcase.get("source_snapshot_version", "")
        requirement_ids = testcase.get("related_requirement_ids") or []
        source_refs = testcase.get("source_refs") or []
        related_change_ids = testcase.get("related_change_ids") or []

        if isinstance(requirement_ids, str):
            requirement_ids = [
                item.strip()
                for item in requirement_ids.split(",")
                if item.strip()
            ]

        if isinstance(source_refs, str):
            source_refs = [
                item.strip()
                for item in source_refs.split(",")
                if item.strip()
            ]

        if isinstance(related_change_ids, str):
            related_change_ids = [
                item.strip()
                for item in related_change_ids.split(",")
                if item.strip()
            ]

        for change_id in related_change_ids:
            change = change_by_id.get(change_id)
            if not change:
                continue
            source_refs.append(
                f"{change.get('source_type', '')}:{change.get('source_id', '')}"
            )

        if not requirement_ids:
            requirement_ids = [""]

        if not source_refs:
            source_refs = [""]

        for source_ref in source_refs:
            source_text = str(source_ref)
            source_type = ""
            source_id = source_text

            if ":" in source_text:
                source_type, source_id = source_text.split(":", 1)

            for requirement_id in requirement_ids:
                ws.append(
                    [
                        source_type,
                        source_id,
                        requirement_id,
                        scenario_id,
                        testcase_id,
                        snapshot_version,
                    ]
                )

    _apply_table_style(ws)
    _auto_width(ws)


def export_function_based_testcases_to_excel(
    ticket_id: str,
    testcases: list,
    coverage_review: dict | None = None,
    final_coverage_review: dict | None = None,
    approved_structure: dict | None = None,
    analysis: dict | None = None,
    clarifications: dict | None = None,
    clarification_answers: dict | None = None,
    requirement_summary: dict | None = None,
) -> str:
    coverage_review = coverage_review or {}
    final_coverage_review = final_coverage_review or {}

    if not approved_structure:
        approved_structure = load_approved_test_case_structure(ticket_id)

    wb = Workbook()

    grouped_testcases = _group_testcases_by_function(
        testcases=testcases,
        approved_structure=approved_structure or {},
    )
    
    analysis = analysis or {}
    clarifications = clarifications or {}
    clarification_answers = clarification_answers or {}
    requirement_summary = requirement_summary or {}

    _create_summary_sheet(
        wb=wb,
        ticket_id=ticket_id,
        approved_structure=approved_structure or {},
        testcases=testcases,
        coverage_review=coverage_review,
        final_coverage_review=final_coverage_review,
    )

    _create_requirements_sheet(
        wb=wb,
        analysis=analysis,
    )

    _create_requirement_matrix_sheet(
        wb=wb,
        analysis=analysis,
        testcases=testcases,
    )

    _create_clarifications_sheet(
        wb=wb,
        clarifications=clarifications,
        clarification_answers=clarification_answers,
    )

    _create_requirement_summary_sheet(
        wb=wb,
        requirement_summary=requirement_summary,
    )

    _create_test_case_structure_sheet(
        wb=wb,
        approved_structure=approved_structure or {},
    )

    _create_master_testcases_sheet(
        wb=wb,
        testcases=testcases,
    )

    _create_function_testcase_sheets(
        wb=wb,
        grouped_testcases=grouped_testcases,
    )

    _create_coverage_review_sheet(
        wb=wb,
        coverage_review=coverage_review,
    )

    _create_final_review_sheet(
        wb=wb,
        final_coverage_review=final_coverage_review,
    )

    _create_traceability_matrix_sheet(
        wb=wb,
        testcases=testcases,
    )

    exports_dir = _ensure_exports_dir(ticket_id)
    export_file = exports_dir / f"{ticket_id}_function_based_testcases.xlsx"

    wb.save(export_file)

    return str(export_file)


def _next_incremental_export_file(ticket_id: str) -> Path:
    exports_dir = _ensure_exports_dir(ticket_id)
    max_version = 0
    pattern = re.compile(rf"{re.escape(ticket_id)}_incremental_testcases_v(\d+)\.xlsx$")

    for path in exports_dir.glob(f"{ticket_id}_incremental_testcases_v*.xlsx"):
        match = pattern.match(path.name)
        if match:
            max_version = max(max_version, int(match.group(1)))

    return exports_dir / f"{ticket_id}_incremental_testcases_v{max_version + 1}.xlsx"


def export_incremental_testcases_to_excel(
    ticket_id: str,
    testcases: list,
    change_impact_report: dict | None = None,
    regeneration_plan: dict | None = None,
    merge_report: dict | None = None,
    coverage_review: dict | None = None,
    final_coverage_review: dict | None = None,
    approved_structure: dict | None = None,
    analysis: dict | None = None,
    clarifications: dict | None = None,
    clarification_answers: dict | None = None,
    requirement_summary: dict | None = None,
) -> str:
    coverage_review = coverage_review or {}
    final_coverage_review = final_coverage_review or {}
    change_impact_report = change_impact_report or {}
    regeneration_plan = regeneration_plan or {}
    merge_report = merge_report or {}

    if not approved_structure:
        approved_structure = load_approved_test_case_structure(ticket_id)

    wb = Workbook()
    grouped_testcases = _group_testcases_by_function(
        testcases=testcases,
        approved_structure=approved_structure or {},
    )

    _create_summary_sheet(
        wb=wb,
        ticket_id=ticket_id,
        approved_structure=approved_structure or {},
        testcases=testcases,
        coverage_review=coverage_review,
        final_coverage_review=final_coverage_review,
    )
    _append_kv(wb["Summary"], "Incremental Export", "Yes")
    _append_kv(wb["Summary"], "Incremental Merge Report", merge_report.get("version", ""))
    _auto_width(wb["Summary"])

    _create_requirements_sheet(wb=wb, analysis=analysis or {})
    _create_requirement_matrix_sheet(wb=wb, analysis=analysis or {}, testcases=testcases)
    _create_clarifications_sheet(
        wb=wb,
        clarifications=clarifications or {},
        clarification_answers=clarification_answers or {},
    )
    _create_requirement_summary_sheet(wb=wb, requirement_summary=requirement_summary or {})
    _create_test_case_structure_sheet(wb=wb, approved_structure=approved_structure or {})
    _create_incremental_testcases_sheet(wb=wb, testcases=testcases)
    _create_master_testcases_sheet(wb=wb, testcases=testcases)
    _create_function_testcase_sheets(wb=wb, grouped_testcases=grouped_testcases)
    _create_coverage_review_sheet(wb=wb, coverage_review=coverage_review)
    _create_final_review_sheet(wb=wb, final_coverage_review=final_coverage_review)
    _create_traceability_matrix_sheet(wb=wb, testcases=testcases)
    _create_change_impact_sheet(wb=wb, change_impact_report=change_impact_report)
    _create_regeneration_plan_sheet(wb=wb, regeneration_plan=regeneration_plan)
    _create_deprecated_replaced_sheet(wb=wb, testcases=testcases)
    _create_source_traceability_sheet(
        wb=wb,
        testcases=testcases,
        change_impact_report=change_impact_report,
    )

    export_file = _next_incremental_export_file(ticket_id)
    wb.save(export_file)

    return str(export_file)
