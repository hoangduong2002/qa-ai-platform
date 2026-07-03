import json
import logging
from pathlib import Path
from datetime import datetime, timezone
from openpyxl import Workbook
from app.services.requirement_list_service import list_requirements
from app.utils.artifact_loader import load_ticket_artifacts
from app.utils.ai_usage_logger import get_usage_log_file
from app.utils.ai_usage_report import normalize_node_name


logger = logging.getLogger(__name__)


def _coalesce(*values):
    for value in values:
        if value not in (None, ""):
            return value
    return ""


def _safe_number(value):
    if value in (None, ""):
        return 0

    try:
        return int(value)
    except (TypeError, ValueError):
        try:
            return float(value)
        except (TypeError, ValueError):
            return 0


def _parse_usage_timestamp(log: dict) -> datetime:
    timestamp = _coalesce(
        log.get("timestamp"),
        log.get("created_at"),
        log.get("time"),
    )

    if not timestamp:
        return datetime.min.replace(tzinfo=timezone.utc)

    try:
        normalized = str(timestamp).replace("Z", "+00:00")
        parsed = datetime.fromisoformat(normalized)
    except ValueError:
        return datetime.min.replace(tzinfo=timezone.utc)

    if parsed.tzinfo is None:
        return parsed.replace(tzinfo=timezone.utc)

    return parsed


def _normalize_usage_log(log: dict) -> dict:
    timestamp = _coalesce(
        log.get("timestamp"),
        log.get("created_at"),
        log.get("time"),
    )
    provider = _coalesce(
        log.get("provider"),
        log.get("llm_provider"),
        log.get("ai_provider"),
    )
    input_tokens = _coalesce(
        log.get("input_tokens"),
        log.get("estimated_input_tokens"),
    )
    output_tokens = _coalesce(
        log.get("output_tokens"),
        log.get("estimated_output_tokens"),
    )
    duration_seconds = log.get("duration_seconds")

    if duration_seconds in (None, "") and log.get("duration_ms") not in (None, ""):
        duration_seconds = _safe_number(log.get("duration_ms")) / 1000

    total_tokens = log.get("total_tokens")

    if total_tokens in (None, ""):
        total_tokens = _safe_number(input_tokens) + _safe_number(output_tokens)

    normalized = dict(log)
    normalized.update(
        {
            "timestamp": timestamp,
            "provider": provider,
            "input_tokens": input_tokens,
            "output_tokens": output_tokens,
            "total_tokens": total_tokens,
            "duration_seconds": duration_seconds,
            "prompt_chars": _coalesce(log.get("prompt_chars"), log.get("input_chars")),
            "response_chars": _coalesce(log.get("response_chars"), log.get("output_chars")),
            "_sort_timestamp": _parse_usage_timestamp(log),
        }
    )

    return normalized


def load_ai_usage_logs():
    log_file = get_usage_log_file()
    logger.info("Loading AI usage logs path=%s", log_file)

    if not log_file.exists():
        logger.info(
            "Loaded AI usage logs path=%s record_count=0 latest_timestamp=",
            log_file,
        )
        return []

    logs = []

    for line_number, line in enumerate(log_file.read_text(encoding="utf-8").splitlines(), start=1):
        if not line.strip():
            continue
        try:
            record = json.loads(line)
        except json.JSONDecodeError:
            logger.warning(
                "Skipping malformed AI usage log line path=%s line=%s",
                log_file,
                line_number,
            )
            continue

        if isinstance(record, dict):
            logs.append(_normalize_usage_log(record))

    logs.sort(key=lambda log: log["_sort_timestamp"], reverse=True)
    latest_timestamp = logs[0].get("timestamp", "") if logs else ""
    logger.info(
        "Loaded AI usage logs path=%s record_count=%s latest_timestamp=%s",
        log_file,
        len(logs),
        latest_timestamp,
    )

    return logs


def generate_system_report():
    output_dir = Path("reports")
    output_dir.mkdir(parents=True, exist_ok=True)

    output_file = (
        output_dir
        / f"qa_ai_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    )

    requirements = list_requirements()
    usage_logs = load_ai_usage_logs()

    wb = Workbook()

    ws = wb.active
    ws.title = "Requirement Summary"

    ws.append(
        [
            "Requirement ID",
            "Name",
            "Created At",
            "Status",
            "Scenario Count",
            "Test Case Count",
            "Improve Iterations",
            "AI Request Count",
            "Input Tokens",
            "Output Tokens",
            "Total Tokens",
            "Total AI Duration Seconds"
        ]
    )

    for item in requirements:
        ticket_id = item["ticket_id"]

        artifacts = load_ticket_artifacts(ticket_id)

        ticket_logs = [
            log for log in usage_logs
            if log.get("ticket_id") == ticket_id
        ]

        input_tokens = sum(
            log.get("input_tokens") or 0
            for log in ticket_logs
        )

        output_tokens = sum(
            log.get("output_tokens") or 0
            for log in ticket_logs
        )

        duration = sum(
            log.get("duration_seconds") or 0
            for log in ticket_logs
        )

        scenarios = artifacts.get("scenarios", [])
        testcases = (
            artifacts.get("improved_testcases")
            or artifacts.get("testcases", [])
        )

        session = artifacts.get("session", {})

        ws.append(
            [
                ticket_id,
                item.get("summary", ""),
                item.get("created_at", ""),
                item.get("status", ""),
                len(scenarios),
                len(testcases),
                session.get("improve_iterations", 0),
                len(ticket_logs),
                input_tokens,
                output_tokens,
                input_tokens + output_tokens,
                round(duration, 2)
            ]
        )

    ws_logs = wb.create_sheet("AI Usage Logs")

    ws_logs.append(
        [
            "Timestamp",
            "Requirement ID",
            "Node",
            "Provider",
            "Model",
            "Input Tokens",
            "Output Tokens",
            "Total Tokens",
            "Duration Seconds",
            "Prompt Chars",
            "Response Chars"
        ]
    )

    for log in usage_logs:
        ws_logs.append(
            [
                log.get("timestamp", ""),
                log.get("ticket_id", ""),
                log.get("node_name", ""),
                log.get("provider", ""),
                log.get("model", ""),
                log.get("input_tokens", ""),
                log.get("output_tokens", ""),
                log.get("total_tokens", ""),
                log.get("duration_seconds", ""),
                log.get("prompt_chars", ""),
                log.get("response_chars", "")
            ]
        )
        
    ws_node = wb.create_sheet(
        "AI Usage by Node"
    )

    ws_node.append(
        [
            "Node",
            "AI Request Count",
            "Input Tokens",
            "Output Tokens",
            "Total Tokens",
            "Total Duration Seconds"
        ]
    )

    node_stats = {}

    for log in usage_logs:
        node = normalize_node_name(
            log.get("node_name", "unknown")
        )

        if node not in node_stats:
            node_stats[node] = {
                "count": 0,
                "input_tokens": 0,
                "output_tokens": 0,
                "duration": 0
            }

        node_stats[node]["count"] += 1
        node_stats[node]["input_tokens"] += (
            log.get("input_tokens") or 0
        )
        node_stats[node]["output_tokens"] += (
            log.get("output_tokens") or 0
        )
        node_stats[node]["duration"] += (
            log.get("duration_seconds") or 0
        )

    for node, stat in sorted(
        node_stats.items(),
        key=lambda x: (
            x[1]["input_tokens"]
            + x[1]["output_tokens"]
        ),
        reverse=True
    ):
        total_tokens = (
            stat["input_tokens"]
            + stat["output_tokens"]
        )

        ws_node.append(
            [
                node,
                stat["count"],
                stat["input_tokens"],
                stat["output_tokens"],
                total_tokens,
                round(stat["duration"], 2)
            ]
        )

    wb.save(output_file)
    
    

    return str(output_file)
