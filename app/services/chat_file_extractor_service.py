import csv
import json
import os
import re
from dataclasses import dataclass, asdict
from pathlib import Path
from typing import Any

from fastapi import UploadFile
from docx import Document
from openpyxl import load_workbook


SUPPORTED_EXTENSIONS = {".txt", ".md", ".json", ".csv", ".xlsx", ".docx", ".pdf"}
DEFAULT_MAX_UPLOAD_MB = 10
DEFAULT_MAX_EXTRACTED_CHARS = 60000


@dataclass
class ExtractedChatFile:
    filename: str
    stored_filename: str
    extension: str
    extracted_chars: int = 0
    warning: str = ""

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)


def _env_int(name: str, default: int) -> int:
    try:
        return int(os.getenv(name, str(default)))
    except ValueError:
        return default


def get_max_upload_bytes() -> int:
    return max(_env_int("CHAT_MAX_UPLOAD_MB", DEFAULT_MAX_UPLOAD_MB), 1) * 1024 * 1024


def get_max_extracted_chars() -> int:
    return max(_env_int("CHAT_MAX_EXTRACTED_CHARS", DEFAULT_MAX_EXTRACTED_CHARS), 1)


def sanitize_upload_filename(filename: str) -> str:
    safe_name = Path(str(filename or "upload")).name
    safe_name = re.sub(r"[^A-Za-z0-9._-]+", "_", safe_name).strip("._")

    if not safe_name:
        safe_name = "upload"

    return safe_name[:180]


def _safe_child_path(parent: Path, filename: str) -> Path:
    parent = parent.resolve()
    candidate = (parent / filename).resolve()

    if parent != candidate.parent:
        raise ValueError("Invalid upload filename.")

    return candidate


async def save_and_extract_uploads(
    session_dir: Path,
    files: list[UploadFile],
) -> tuple[list[dict[str, Any]], str, list[str]]:
    original_dir = session_dir / "uploads" / "original"
    extracted_dir = session_dir / "uploads" / "extracted"
    original_dir.mkdir(parents=True, exist_ok=True)
    extracted_dir.mkdir(parents=True, exist_ok=True)

    attachments: list[dict[str, Any]] = []
    context_parts: list[str] = []
    warnings: list[str] = []
    remaining_context_chars = get_max_extracted_chars()

    for upload in files or []:
        if not upload or not upload.filename:
            continue

        result, text = await save_and_extract_upload(
            upload=upload,
            original_dir=original_dir,
            extracted_dir=extracted_dir,
            remaining_context_chars=remaining_context_chars,
        )
        result_dict = result.to_dict()
        attachments.append(result_dict)

        if result.warning:
            warnings.append(result.warning)

        if text:
            context_parts.append(f"## File: {result.filename}\n{text}")
            remaining_context_chars = max(remaining_context_chars - len(text), 0)

    return attachments, "\n\n".join(context_parts), warnings


async def save_and_extract_upload(
    upload: UploadFile,
    original_dir: Path,
    extracted_dir: Path,
    remaining_context_chars: int,
) -> tuple[ExtractedChatFile, str]:
    filename = sanitize_upload_filename(upload.filename)
    extension = Path(filename).suffix.lower()
    stored_path = _safe_child_path(original_dir, filename)

    if extension not in SUPPORTED_EXTENSIONS:
        warning = f"Unsupported file type: {filename}"
        return ExtractedChatFile(filename, filename, extension, warning=warning), ""

    max_upload_bytes = get_max_upload_bytes()
    size = 0

    with stored_path.open("wb") as output:
        while True:
            chunk = await upload.read(1024 * 1024)
            if not chunk:
                break
            size += len(chunk)
            if size > max_upload_bytes:
                stored_path.unlink(missing_ok=True)
                warning = f"File is too large: {filename}"
                return ExtractedChatFile(filename, filename, extension, warning=warning), ""
            output.write(chunk)

    try:
        extracted_text = extract_text_from_file(stored_path)
    except Exception:
        warning = f"Could not extract text from {filename}."
        return ExtractedChatFile(filename, filename, extension, warning=warning), ""

    extracted_text = (extracted_text or "").strip()
    if not extracted_text:
        warning = f"No text could be extracted from {filename}."
        return ExtractedChatFile(filename, filename, extension, warning=warning), ""

    limited_text = extracted_text[: max(remaining_context_chars, 0)]
    extracted_path = _safe_child_path(extracted_dir, f"{stored_path.stem}.txt")
    extracted_path.write_text(limited_text, encoding="utf-8")

    return (
        ExtractedChatFile(
            filename=filename,
            stored_filename=stored_path.name,
            extension=extension,
            extracted_chars=len(limited_text),
        ),
        limited_text,
    )


def extract_text_from_file(file_path: Path) -> str:
    suffix = file_path.suffix.lower()

    if suffix in {".txt", ".md"}:
        return file_path.read_text(encoding="utf-8", errors="ignore")

    if suffix == ".json":
        data = json.loads(file_path.read_text(encoding="utf-8", errors="ignore"))
        return json.dumps(data, indent=2, ensure_ascii=False)

    if suffix == ".csv":
        return _extract_csv_text(file_path)

    if suffix == ".xlsx":
        return _extract_xlsx_text(file_path)

    if suffix == ".docx":
        return _extract_docx_text(file_path)

    if suffix == ".pdf":
        return _extract_pdf_text(file_path)

    raise ValueError(f"Unsupported file type: {file_path.name}")


def _extract_csv_text(file_path: Path) -> str:
    rows: list[str] = []

    with file_path.open("r", encoding="utf-8", errors="ignore", newline="") as handle:
        for row in csv.reader(handle):
            rows.append(", ".join(row))

    return "\n".join(rows)


def _extract_xlsx_text(file_path: Path) -> str:
    workbook = load_workbook(file_path, read_only=True, data_only=True)
    parts: list[str] = []

    for sheet in workbook.worksheets:
        parts.append(f"# Sheet: {sheet.title}")
        for row in sheet.iter_rows(values_only=True):
            values = [str(value) for value in row if value is not None]
            if values:
                parts.append(", ".join(values))

    workbook.close()
    return "\n".join(parts)


def _extract_docx_text(file_path: Path) -> str:
    document = Document(str(file_path))
    return "\n".join(
        paragraph.text
        for paragraph in document.paragraphs
        if paragraph.text.strip()
    )


def _extract_pdf_text(file_path: Path) -> str:
    try:
        from pypdf import PdfReader
    except ImportError as error:
        raise RuntimeError("PDF extraction requires pypdf.") from error

    reader = PdfReader(str(file_path))
    return "\n".join(
        page.extract_text() or ""
        for page in reader.pages
    )
