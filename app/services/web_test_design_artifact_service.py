import json
import re
from pathlib import Path
from typing import Any

from app.exporters.function_based_excel_exporter import (
    export_incremental_testcases_to_excel,
    export_function_based_testcases_to_excel,
)
from app.services.llm_router_service import (
    TASK_SCENARIO_COVERAGE_REVIEW,
    TASK_SCENARIO_IMPROVEMENT,
    call_text_llm,
)
from app.services.portal_ai_mode_service import get_current_portal_ai_mode
from app.utils.artifact_loader import load_ticket_artifacts
from app.utils.llm_json import parse_json
from app.utils.test_structure_store import load_approved_test_case_structure
from graph.nodes.final_review_coverage import final_coverage_review
from graph.nodes.improve_testcases import improve_testcases
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


def _current_ai_mode() -> str | None:
    portal_ai_mode = get_current_portal_ai_mode()
    if portal_ai_mode:
        return portal_ai_mode.get("ai_mode")

    return None


def _resolve_ai_mode(ai_mode: str | None = None) -> str | None:
    return ai_mode or _current_ai_mode()


def _apply_web_ai_state(state: dict, ai_mode: str | None = None) -> dict:
    resolved_ai_mode = _resolve_ai_mode(ai_mode)

    if resolved_ai_mode:
        state["ai_mode"] = resolved_ai_mode

    state["source_channel"] = "web"
    return state


def _read_json(file_path: Path, default: Any):
    if not file_path.exists():
        return default

    return json.loads(file_path.read_text(encoding="utf-8"))


def _write_json(file_path: Path, data: Any) -> str:
    file_path.parent.mkdir(parents=True, exist_ok=True)
    file_path.write_text(
        json.dumps(data, indent=2, ensure_ascii=False),
        encoding="utf-8",
    )
    return str(file_path)


def _root(ticket_id: str) -> Path:
    return Path("requirements") / ticket_id


def _scenario_dir(ticket_id: str) -> Path:
    return _root(ticket_id) / "scenarios"


def _testcase_dir(ticket_id: str) -> Path:
    return _root(ticket_id) / "testcases"


def _review_dir(ticket_id: str) -> Path:
    return _root(ticket_id) / "review"


def _scenario_session_file(ticket_id: str) -> Path:
    return _scenario_dir(ticket_id) / "scenario_session.json"


def _testcase_session_file(ticket_id: str) -> Path:
    return _testcase_dir(ticket_id) / "testcase_session.json"


def _version_number(version: str) -> int:
    if re.match(r"v\d+$", version or ""):
        return int(version.replace("v", ""))
    return 0


def _next_version(existing_versions: list[str]) -> str:
    max_number = 0

    for version in existing_versions:
        max_number = max(max_number, _version_number(version))

    return f"v{max_number + 1}"


def _default_scenario_session() -> dict:
    return {
        "current_version": None,
        "approved": False,
        "approved_version": None,
        "last_review_version": None,
    }


def _default_testcase_session() -> dict:
    return {
        "current_version": None,
        "approved": False,
        "approved_version": None,
        "last_final_review_version": None,
    }


def load_scenario_session(ticket_id: str) -> dict:
    session = _default_scenario_session()
    session.update(_read_json(_scenario_session_file(ticket_id), {}) or {})
    return session


def save_scenario_session(ticket_id: str, session: dict) -> str:
    merged = _default_scenario_session()
    merged.update(session or {})
    return _write_json(_scenario_session_file(ticket_id), merged)


def load_testcase_session(ticket_id: str) -> dict:
    session = _default_testcase_session()
    session.update(_read_json(_testcase_session_file(ticket_id), {}) or {})
    return session


def save_testcase_session(ticket_id: str, session: dict) -> str:
    merged = _default_testcase_session()
    merged.update(session or {})
    return _write_json(_testcase_session_file(ticket_id), merged)


def list_scenario_versions(ticket_id: str) -> list[dict]:
    scenario_dir = _scenario_dir(ticket_id)
    versions = []

    if (scenario_dir / "scenarios.json").exists():
        versions.append({"version": "latest", "label": "Latest Draft"})

    if (scenario_dir / "approved_scenarios.json").exists():
        versions.append({"version": "approved", "label": "Approved"})

    if scenario_dir.exists():
        files = sorted(
            scenario_dir.glob("scenarios_v*.json"),
            key=lambda path: _version_number(
                path.stem.replace("scenarios_", "")
            ),
        )

        for file_path in files:
            version = file_path.stem.replace("scenarios_", "")
            versions.append({"version": version, "label": version})

    return versions


def list_testcase_versions(ticket_id: str) -> list[dict]:
    testcase_dir = _testcase_dir(ticket_id)
    versions = []

    if (testcase_dir / "testcases.json").exists():
        versions.append({"version": "latest", "label": "Latest Draft"})

    if (testcase_dir / "approved_testcases.json").exists():
        versions.append({"version": "approved", "label": "Approved"})

    if testcase_dir.exists():
        files = sorted(
            testcase_dir.glob("testcases_v*.json"),
            key=lambda path: _version_number(
                path.stem.replace("testcases_", "")
            ),
        )

        for file_path in files:
            version = file_path.stem.replace("testcases_", "")
            versions.append({"version": version, "label": version})

    return versions


def _scenario_path(ticket_id: str, version: str) -> Path:
    if version == "approved":
        return _scenario_dir(ticket_id) / "approved_scenarios.json"
    if version == "latest":
        return _scenario_dir(ticket_id) / "scenarios.json"
    return _scenario_dir(ticket_id) / f"scenarios_{version}.json"


def _test_scope_path(ticket_id: str, version: str) -> Path:
    if version == "latest":
        return _scenario_dir(ticket_id) / "test_scope.json"
    return _scenario_dir(ticket_id) / f"test_scope_{version}.json"


def _coverage_review_path(ticket_id: str, version: str) -> Path:
    if version == "latest":
        version = load_scenario_session(ticket_id).get("current_version") or "v1"
    return _scenario_dir(ticket_id) / f"coverage_review_{version}.json"


def _testcase_path(ticket_id: str, version: str) -> Path:
    if version == "approved":
        return _testcase_dir(ticket_id) / "approved_testcases.json"
    if version == "latest":
        return _testcase_dir(ticket_id) / "testcases.json"
    return _testcase_dir(ticket_id) / f"testcases_{version}.json"


def _final_review_path(ticket_id: str, version: str) -> Path:
    if version == "latest":
        version = load_testcase_session(ticket_id).get("current_version") or "v1"
    return _testcase_dir(ticket_id) / f"final_review_{version}.json"


def get_scenarios(ticket_id: str, version: str = "latest") -> list:
    return _read_json(_scenario_path(ticket_id, version), [])


def get_scenarios_json(ticket_id: str, version: str = "latest") -> str:
    data = get_scenarios(ticket_id, version)
    return json.dumps(data, indent=2, ensure_ascii=False) if data else ""


def get_test_scope(ticket_id: str, version: str = "latest") -> dict:
    return _read_json(_test_scope_path(ticket_id, version), {})


def get_coverage_review(ticket_id: str, version: str = "latest") -> dict:
    return _read_json(_coverage_review_path(ticket_id, version), {})


def get_coverage_review_json(ticket_id: str, version: str = "latest") -> str:
    data = get_coverage_review(ticket_id, version)
    return json.dumps(data, indent=2, ensure_ascii=False) if data else ""


def get_testcases(ticket_id: str, version: str = "latest") -> list:
    return _read_json(_testcase_path(ticket_id, version), [])


def get_testcases_json(ticket_id: str, version: str = "latest") -> str:
    data = get_testcases(ticket_id, version)
    return json.dumps(data, indent=2, ensure_ascii=False) if data else ""


def get_final_review(ticket_id: str, version: str = "latest") -> dict:
    return _read_json(_final_review_path(ticket_id, version), {})


def get_final_review_json(ticket_id: str, version: str = "latest") -> str:
    data = get_final_review(ticket_id, version)
    return json.dumps(data, indent=2, ensure_ascii=False) if data else ""


def _coverage_prompt(state: dict, scenarios: list, test_scope: dict) -> str:
    return f"""
Review scenario coverage. Return ONLY JSON object.

Focus on Requirement -> Scope -> Scenario coverage.
Do not review test case quality.

JSON schema:
{{
  "coverage_score": 0,
  "approved_by_ai": false,
  "summary": "",
  "missing_scenarios": [],
  "weak_scenarios": [],
  "duplicate_scenarios": [],
  "recommendations": []
}}

Requirement summary:
{json.dumps(state.get("requirement_summary", {}), ensure_ascii=False)}

Requirement items:
{json.dumps(state.get("analysis", {}).get("requirement_items", []), ensure_ascii=False)}

Test scope:
{json.dumps(test_scope, ensure_ascii=False)}

Scenarios:
{json.dumps(scenarios, ensure_ascii=False)}
"""


def run_scenario_coverage_review(
    ticket_id: str,
    version: str,
    ai_mode: str | None = None,
) -> dict:
    scenarios = get_scenarios(ticket_id, version)
    if not scenarios:
        raise ValueError("Scenarios are required before coverage review.")

    test_scope = get_test_scope(ticket_id, version)
    state = load_ticket_artifacts(ticket_id)
    ai_mode = _resolve_ai_mode(ai_mode)

    response_content = call_text_llm(
        TASK_SCENARIO_COVERAGE_REVIEW,
        _coverage_prompt(state, scenarios, test_scope),
        ai_mode=ai_mode,
        source_channel="web",
    )

    review = parse_json(response_content)
    if not isinstance(review, dict):
        raise ValueError("Coverage review must be a JSON object.")

    review_version = (
        load_scenario_session(ticket_id).get("current_version")
        if version == "latest"
        else version
    ) or "v1"

    _write_json(_coverage_review_path(ticket_id, review_version), review)
    _write_json(_review_dir(ticket_id) / "coverage_review.json", review)

    session = load_scenario_session(ticket_id)
    session["last_review_version"] = review_version
    save_scenario_session(ticket_id, session)

    return review


def _review_to_comment(review: dict) -> str:
    return json.dumps(review, indent=2, ensure_ascii=False)


def _normalize_scenarios(data: Any) -> list:
    if isinstance(data, list):
        return data

    if isinstance(data, dict):
        for key in ["scenarios", "test_scenarios", "testScenarios"]:
            if isinstance(data.get(key), list):
                return data[key]

    raise ValueError("Invalid scenarios JSON.")


def _improve_scenarios(
    ticket_id: str,
    version: str,
    comment: str,
    ai_mode: str | None = None,
) -> str:
    comment = (comment or "").strip()
    if not comment:
        raise ValueError("Improve comment is required.")

    scenarios = get_scenarios(ticket_id, version)
    test_scope = get_test_scope(ticket_id, version)
    state = load_ticket_artifacts(ticket_id)

    prompt = f"""
Improve scenarios from review comment.
Return ONLY JSON array of scenarios.
Keep rule: 1 scenario = 1 test case.

Review comment:
{comment}

Requirement summary:
{json.dumps(state.get("requirement_summary", {}), ensure_ascii=False)}

Test scope:
{json.dumps(test_scope, ensure_ascii=False)}

Current scenarios:
{json.dumps(scenarios, ensure_ascii=False)}
"""

    response_content = call_text_llm(
        TASK_SCENARIO_IMPROVEMENT,
        prompt,
        ai_mode=_resolve_ai_mode(ai_mode),
        source_channel="web",
    )

    improved_scenarios = _normalize_scenarios(parse_json(response_content))

    existing = [
        item["version"]
        for item in list_scenario_versions(ticket_id)
        if re.match(r"v\d+$", item["version"])
    ]
    new_version = _next_version(existing)

    _write_json(_scenario_path(ticket_id, new_version), improved_scenarios)
    _write_json(_scenario_path(ticket_id, "latest"), improved_scenarios)
    _write_json(_root(ticket_id) / "analysis" / "scenarios.json", improved_scenarios)

    session = load_scenario_session(ticket_id)
    session["current_version"] = new_version
    session["approved"] = False
    save_scenario_session(ticket_id, session)

    return new_version


def improve_scenarios_from_ai_review(
    ticket_id: str,
    version: str,
    ai_mode: str | None = None,
) -> str:
    review = get_coverage_review(ticket_id, version)
    if not review:
        raise ValueError("No scenario coverage review found.")

    return _improve_scenarios(
        ticket_id,
        version,
        _review_to_comment(review),
        ai_mode=ai_mode,
    )


def improve_scenarios_from_human_review(
    ticket_id: str,
    version: str,
    comment: str,
    ai_mode: str | None = None,
) -> str:
    return _improve_scenarios(ticket_id, version, comment, ai_mode=ai_mode)


def approve_scenarios(ticket_id: str, version: str) -> list:
    scenarios = get_scenarios(ticket_id, version)
    if not scenarios:
        raise ValueError("No scenarios found.")

    _write_json(_scenario_path(ticket_id, "approved"), scenarios)

    session = load_scenario_session(ticket_id)
    session["approved"] = True
    session["approved_version"] = version
    save_scenario_session(ticket_id, session)

    return scenarios


def run_final_review(
    ticket_id: str,
    version: str,
    ai_mode: str | None = None,
) -> dict:
    testcases = get_testcases(ticket_id, version)
    scenarios = get_scenarios(ticket_id, "approved")

    if not testcases:
        raise ValueError("Test cases are required before final review.")

    state = load_ticket_artifacts(ticket_id)
    state["ticket_id"] = ticket_id
    state["approved_test_case_structure"] = load_approved_test_case_structure(ticket_id)
    state["scenarios"] = scenarios
    state["testcases"] = testcases
    state["coverage_review"] = get_coverage_review(ticket_id, "latest")
    _apply_web_ai_state(state, ai_mode)

    result = final_coverage_review(state)
    review = result.get("final_coverage_review", {})

    review_version = (
        load_testcase_session(ticket_id).get("current_version")
        if version == "latest"
        else version
    ) or "v1"

    _write_json(_final_review_path(ticket_id, review_version), review)
    _write_json(_review_dir(ticket_id) / "final_coverage_review.json", review)

    session = load_testcase_session(ticket_id)
    session["last_final_review_version"] = review_version
    save_testcase_session(ticket_id, session)

    return review


def _improve_testcases(
    ticket_id: str,
    version: str,
    comment: str,
    ai_mode: str | None = None,
) -> str:
    comment = (comment or "").strip()
    if not comment:
        raise ValueError("Improve comment is required.")

    testcases = get_testcases(ticket_id, version)
    scenarios = get_scenarios(ticket_id, "approved")

    if not testcases:
        raise ValueError("Test cases are required before improve.")

    state = load_ticket_artifacts(ticket_id)
    state["ticket_id"] = ticket_id
    state["approved_test_case_structure"] = load_approved_test_case_structure(ticket_id)
    state["scenarios"] = scenarios
    state["testcases"] = testcases
    state["coverage_review"] = get_final_review(ticket_id, version) or {}
    state["review_comments"] = [comment]
    state["improve_version"] = "web"
    _apply_web_ai_state(state, ai_mode)

    result = improve_testcases(state)
    improved = result.get("improved_testcases") or result.get("testcases")

    if not improved:
        raise ValueError("Improve test cases failed.")

    existing = [
        item["version"]
        for item in list_testcase_versions(ticket_id)
        if re.match(r"v\d+$", item["version"])
    ]
    new_version = _next_version(existing)

    _write_json(_testcase_path(ticket_id, new_version), improved)
    _write_json(_testcase_path(ticket_id, "latest"), improved)

    session = load_testcase_session(ticket_id)
    session["current_version"] = new_version
    session["approved"] = False
    save_testcase_session(ticket_id, session)

    return new_version


def improve_testcases_from_ai_review(
    ticket_id: str,
    version: str,
    ai_mode: str | None = None,
) -> str:
    review = get_final_review(ticket_id, version)
    if not review:
        raise ValueError("No final review found.")

    return _improve_testcases(
        ticket_id,
        version,
        _review_to_comment(review),
        ai_mode=ai_mode,
    )


def improve_testcases_from_human_review(
    ticket_id: str,
    version: str,
    comment: str,
    ai_mode: str | None = None,
) -> str:
    return _improve_testcases(ticket_id, version, comment, ai_mode=ai_mode)


def approve_testcases(ticket_id: str, version: str) -> list:
    testcases = get_testcases(ticket_id, version)
    if not testcases:
        raise ValueError("No test cases found.")

    _write_json(_testcase_path(ticket_id, "approved"), testcases)

    session = load_testcase_session(ticket_id)
    session["approved"] = True
    session["approved_version"] = version
    save_testcase_session(ticket_id, session)

    return testcases


def save_testcases_json_as_new_version(ticket_id: str, testcases_json: str) -> str:
    try:
        testcases = json.loads(testcases_json)
    except json.JSONDecodeError as error:
        raise ValueError(f"Invalid JSON: {error}") from error

    if not isinstance(testcases, list):
        raise ValueError("Testcases JSON must be an array.")

    existing = [
        item["version"]
        for item in list_testcase_versions(ticket_id)
        if re.match(r"v\d+$", item["version"])
    ]
    version = _next_version(existing)

    _write_json(_testcase_path(ticket_id, version), testcases)
    _write_json(_testcase_path(ticket_id, "latest"), testcases)

    session = load_testcase_session(ticket_id)
    session["current_version"] = version
    session["approved"] = False
    save_testcase_session(ticket_id, session)

    return version


def export_testcases_excel(ticket_id: str, version: str) -> Path:
    testcases = get_testcases(ticket_id, version)

    if not testcases:
        raise ValueError("No test cases found.")

    excel_file = export_function_based_testcases_to_excel(
        ticket_id=ticket_id,
        testcases=testcases,
        coverage_review=get_coverage_review(ticket_id, "latest"),
        final_coverage_review=get_final_review(ticket_id, version),
        approved_structure=load_approved_test_case_structure(ticket_id),
        analysis=_load_requirement_analysis_for_export(ticket_id),
        clarifications=_load_clarifications_for_export(ticket_id),
        clarification_answers=_load_clarification_answers_for_export(ticket_id),
        requirement_summary=_load_requirement_summary_for_export(ticket_id),
    )

    return Path(excel_file)


def _latest_versioned_json(
    directory: Path,
    pattern: str,
    version_pattern: str,
    default,
):
    if not directory.exists():
        return default

    latest_path = None
    latest_version = 0

    for path in directory.glob(pattern):
        match = re.match(version_pattern, path.name)
        if match and int(match.group(1)) > latest_version:
            latest_version = int(match.group(1))
            latest_path = path

    if latest_path is None:
        return default

    return _read_json_file(latest_path, default=default)


def export_incremental_testcases_excel(ticket_id: str) -> Path:
    root = _root(ticket_id)
    generated_dir = root / "generated"
    analysis_dir = root / "analysis"

    testcases = _read_json_file(
        generated_dir / "latest_testcases.json",
        default=[],
    )

    if not testcases:
        testcases = _latest_versioned_json(
            generated_dir,
            "incremental_testcases_v*.json",
            r"incremental_testcases_v(\d+)\.json$",
            [],
        )

    if not testcases:
        raise ValueError("No incremental test cases found.")

    merge_report = _latest_versioned_json(
        analysis_dir,
        "incremental_testcase_merge_report_v*.json",
        r"incremental_testcase_merge_report_v(\d+)\.json$",
        {},
    )
    change_impact_report = _read_json_file(
        analysis_dir / "latest_change_impact_report.json",
        default={},
    )
    regeneration_plan = _read_json_file(
        analysis_dir / "latest_regeneration_plan.json",
        default={},
    )

    excel_file = export_incremental_testcases_to_excel(
        ticket_id=ticket_id,
        testcases=testcases,
        change_impact_report=change_impact_report,
        regeneration_plan=regeneration_plan,
        merge_report=merge_report,
        coverage_review=get_coverage_review(ticket_id, "latest"),
        final_coverage_review=get_final_review(ticket_id, "latest"),
        approved_structure=load_approved_test_case_structure(ticket_id),
        analysis=_load_requirement_analysis_for_export(ticket_id),
        clarifications=_load_clarifications_for_export(ticket_id),
        clarification_answers=_load_clarification_answers_for_export(ticket_id),
        requirement_summary=_load_requirement_summary_for_export(ticket_id),
    )

    return Path(excel_file)


def _read_text_file(path: Path) -> str:
    if not path.exists():
        return ""

    return path.read_text(
        encoding="utf-8",
        errors="ignore",
    )


def _read_json_file(path: Path, default=None):
    if not path.exists():
        return default

    try:
        return json.loads(
            path.read_text(
                encoding="utf-8",
                errors="ignore",
            )
        )
    except Exception:
        return default
    
    
def _load_requirement_analysis_for_export(ticket_id: str) -> dict:
    root = _root(ticket_id)
    analysis_dir = root / "analysis"

    candidates = [
        analysis_dir / "requirement_analysis.json",
        analysis_dir / "analysis.json",
    ]

    for file_path in candidates:
        data = _read_json_file(file_path, default={})

        if isinstance(data, dict) and data:
            return data

    return {}


def _load_clarifications_for_export(ticket_id: str) -> dict:
    root = _root(ticket_id)
    analysis_dir = root / "analysis"

    candidates = [
        analysis_dir / "clarifications.json",
        analysis_dir / "requirement_clarifications.json",
    ]

    for file_path in candidates:
        data = _read_json_file(file_path, default={})

        if isinstance(data, dict) and data:
            return data

        if isinstance(data, list) and data:
            return {
                "questions": data,
            }

    return {}


def _load_clarification_answers_for_export(ticket_id: str) -> dict:
    root = _root(ticket_id)

    candidates = [
        root / "clarification_answers.json",
        root / "analysis" / "clarification_answers.json",
    ]

    for file_path in candidates:
        data = _read_json_file(file_path, default={})

        if isinstance(data, dict) and data:
            return data

        if isinstance(data, list) and data:
            return {
                "answers": data,
            }

    return {}


def _load_requirement_summary_for_export(ticket_id: str) -> dict:
    root = _root(ticket_id)
    analysis_dir = root / "analysis"

    candidates = [
        analysis_dir / "requirement_summary.json",
        analysis_dir / "summary.json",
    ]

    for file_path in candidates:
        data = _read_json_file(file_path, default={})

        if isinstance(data, dict) and data:
            return data

    return {}


def _load_requirement_rows_for_excel(ticket_id: str) -> list[dict]:
    root = _root(ticket_id)
    analysis_dir = root / "analysis"
    source_dir = root / "source"

    rows = []

    analysis = _read_json_file(
        analysis_dir / "requirement_analysis.json",
        default={},
    ) or {}

    requirement_items = (
        analysis.get("requirements")
        or analysis.get("requirement_items")
        or analysis.get("items")
        or []
    )

    if isinstance(requirement_items, list) and requirement_items:
        for index, item in enumerate(requirement_items, start=1):
            if not isinstance(item, dict):
                rows.append(
                    {
                        "requirement_id": f"REQ-{index:03d}",
                        "title": "",
                        "description": str(item),
                        "priority": "",
                        "source": "requirement_analysis",
                    }
                )
                continue

            rows.append(
                {
                    "requirement_id": (
                        item.get("requirement_id")
                        or item.get("id")
                        or item.get("req_id")
                        or f"REQ-{index:03d}"
                    ),
                    "title": (
                        item.get("title")
                        or item.get("name")
                        or item.get("summary")
                        or ""
                    ),
                    "description": (
                        item.get("description")
                        or item.get("detail")
                        or item.get("content")
                        or item.get("requirement")
                        or ""
                    ),
                    "priority": item.get("priority") or "",
                    "source": "requirement_analysis",
                }
            )

        return rows

    summary = _read_json_file(
        analysis_dir / "requirement_summary.json",
        default={},
    ) or {}

    if isinstance(summary, dict) and summary:
        summary_text = (
            summary.get("summary")
            or summary.get("requirement_summary")
            or summary.get("overview")
            or summary.get("business_summary")
            or ""
        )

        if summary_text:
            rows.append(
                {
                    "requirement_id": "REQ-SUMMARY",
                    "title": "Requirement Summary",
                    "description": summary_text,
                    "priority": "",
                    "source": "requirement_summary",
                }
            )

        key_points = (
            summary.get("key_points")
            or summary.get("main_points")
            or summary.get("requirements")
            or []
        )

        if isinstance(key_points, list):
            for index, item in enumerate(key_points, start=1):
                rows.append(
                    {
                        "requirement_id": f"REQ-SUM-{index:03d}",
                        "title": f"Summary Point {index}",
                        "description": _json_text(item),
                        "priority": "",
                        "source": "requirement_summary",
                    }
                )

        if rows:
            return rows

    sanitized_text = _read_text_file(
        analysis_dir / "sanitized_requirement.md",
    )

    if sanitized_text.strip():
        return [
            {
                "requirement_id": "REQ-RAW",
                "title": "Sanitized Requirement",
                "description": sanitized_text.strip(),
                "priority": "",
                "source": "sanitized_requirement",
            }
        ]

    source_candidates = []

    source_candidates.extend(
        sorted((source_dir / "jira").glob("*.md"))
        if (source_dir / "jira").exists()
        else []
    )

    source_candidates.extend(
        [
            source_dir / "requirement.md",
            source_dir / "raw_requirement.md",
            source_dir / "input.md",
        ]
    )

    for source_file in source_candidates:
        source_text = _read_text_file(source_file)

        if source_text.strip():
            return [
                {
                    "requirement_id": "REQ-SOURCE",
                    "title": source_file.name,
                    "description": source_text.strip(),
                    "priority": "",
                    "source": str(source_file),
                }
            ]

    return []


def _autosize_columns(ws):
    for column_cells in ws.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter

        for cell in column_cells:
            value = cell.value
            if value is None:
                continue

            max_length = max(max_length, len(str(value)))

        ws.column_dimensions[column_letter].width = min(max_length + 2, 60)


def _style_header(ws):
    header_fill = PatternFill(
        start_color="1F2937",
        end_color="1F2937",
        fill_type="solid",
    )

    header_font = Font(
        color="FFFFFF",
        bold=True,
    )

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(
            vertical="center",
            wrap_text=True,
        )


def _json_text(value: Any) -> str:
    if value is None:
        return ""

    if isinstance(value, (dict, list)):
        return json.dumps(
            value,
            ensure_ascii=False,
            indent=2,
        )

    return str(value)


def export_scenarios_excel(
    ticket_id: str,
    version: str,
) -> Path:
    scenarios = get_scenarios(
        ticket_id=ticket_id,
        version=version,
    )

    if not scenarios:
        raise ValueError("No scenarios found.")

    test_scope = get_test_scope(
        ticket_id=ticket_id,
        version=version,
    )

    coverage_review = get_coverage_review(
        ticket_id=ticket_id,
        version=version,
    )

    export_dir = _root(ticket_id) / "exports"
    export_dir.mkdir(parents=True, exist_ok=True)

    safe_version = version.replace("/", "_")
    excel_file = export_dir / f"scenarios_{safe_version}.xlsx"

    wb = Workbook()

    ws = wb.active
    ws.title = "Scenarios"

    ws.append(
        [
            "Scenario ID",
            "Title",
            "Description",
            "Function ID",
            "Sub Function ID",
            "Test Area ID",
            "Scenario Type",
            "Priority",
            "Related Requirement IDs",
            "Preconditions",
            "Test Data",
            "Expected Result",
            "Raw JSON",
        ]
    )

    for index, scenario in enumerate(scenarios, start=1):
        if not isinstance(scenario, dict):
            ws.append(
                [
                    f"SC{index:03d}",
                    "",
                    str(scenario),
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    _json_text(scenario),
                ]
            )
            continue

        ws.append(
            [
                scenario.get("scenario_id")
                or scenario.get("id")
                or f"SC{index:03d}",
                scenario.get("title") or scenario.get("name") or "",
                scenario.get("description") or scenario.get("summary") or "",
                scenario.get("function_id") or "",
                scenario.get("sub_function_id") or "",
                scenario.get("test_area_id") or scenario.get("category_id") or "",
                scenario.get("scenario_type")
                or scenario.get("type")
                or scenario.get("test_type")
                or "",
                scenario.get("priority") or "",
                ", ".join(
                    scenario.get("related_requirement_ids")
                    or scenario.get("related_requirements")
                    or []
                ),
                _json_text(scenario.get("preconditions") or ""),
                _json_text(scenario.get("test_data") or ""),
                _json_text(
                    scenario.get("expected_result")
                    or scenario.get("expected_results")
                    or ""
                ),
                _json_text(scenario),
            ]
        )

    _style_header(ws)
    _autosize_columns(ws)

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )

    ws_scope = wb.create_sheet("Test Scope")
    ws_scope.append(["Field", "Value"])

    if isinstance(test_scope, dict):
        for key, value in test_scope.items():
            ws_scope.append([key, _json_text(value)])
    else:
        ws_scope.append(["test_scope", _json_text(test_scope)])

    _style_header(ws_scope)
    _autosize_columns(ws_scope)

    for row in ws_scope.iter_rows():
        for cell in row:
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )

    ws_review = wb.create_sheet("Coverage Review")
    ws_review.append(["Field", "Value"])

    if isinstance(coverage_review, dict) and coverage_review:
        for key, value in coverage_review.items():
            ws_review.append([key, _json_text(value)])
    else:
        ws_review.append(["coverage_review", "No coverage review found."])

    _style_header(ws_review)
    _autosize_columns(ws_review)

    for row in ws_review.iter_rows():
        for cell in row:
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )

    wb.save(excel_file)

    return excel_file
