import logging
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import (
    Font,
    PatternFill,
    Alignment,
    Border,
    Side
)

from app.utils.clarification_answers import (
    get_clarification_answer_text,
    get_clarification_id,
    get_clarification_question,
    normalize_clarification_answers,
    normalize_question_text,
)


logger = logging.getLogger(__name__)

def _to_text(value):

    if value is None:
        return ""

    if isinstance(value, str):
        return value

    if isinstance(value, (int, float, bool)):
        return str(value)

    if isinstance(value, dict):
        return "\n".join(
            [
                f"{key}: {_to_text(val)}"
                for key, val in value.items()
            ]
        )

    if isinstance(value, list):
        return "\n".join(
            [
                _to_text(item)
                for item in value
            ]
        )

    return str(value)


def _build_requirement_coverage_matrix(
    analysis: dict,
    scenarios: list,
    testcases: list
):
    requirement_items = analysis.get(
        "requirement_items",
        []
    )

    matrix = []

    for item in requirement_items:
        req_id = item.get(
            "requirement_id",
            ""
        )

        related_scenarios = [
            scenario
            for scenario in scenarios
            if req_id in scenario.get(
                "related_requirement_ids",
                []
            )
        ]

        related_scenario_ids = [
            scenario.get(
                "scenario_id",
                ""
            )
            for scenario in related_scenarios
        ]

        related_testcases = [
            testcase
            for testcase in testcases
            if req_id in testcase.get(
                "related_requirement_ids",
                []
            )
        ]

        related_testcase_ids = [
            testcase.get(
                "testcase_id",
                ""
            )
            for testcase in related_testcases
        ]

        matrix.append(
            {
                "requirement_id": req_id,
                "type": item.get(
                    "type",
                    ""
                ),
                "description": item.get(
                    "description",
                    ""
                ),
                "covered": (
                    "Yes"
                    if related_testcases
                    else "No"
                ),
                "scenario_count": len(
                    related_scenarios
                ),
                "testcase_count": len(
                    related_testcases
                ),
                "scenario_ids": related_scenario_ids,
                "testcase_ids": related_testcase_ids
            }
        )

    return matrix


def _build_requirement_lookup(
    analysis: dict
):

    requirement_items = analysis.get(
        "requirement_items",
        []
    )

    return {
        item.get("requirement_id"): item
        for item in requirement_items
        if item.get("requirement_id")
    }


def _format_requirement_details(
    requirement_ids: list,
    requirement_lookup: dict
):

    details = []

    for req_id in requirement_ids:
        item = requirement_lookup.get(req_id)

        if item:
            details.append(
                f"{req_id} - "
                f"{item.get('type', '')}: "
                f"{item.get('description', '')}"
            )
        else:
            details.append(req_id)

    return _to_text(details)


def _apply_styles(wb: Workbook):

    header_fill = PatternFill(
        "solid",
        fgColor="D9EAF7"
    )

    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    for sheet in wb.worksheets:

        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(
                vertical="center",
                wrap_text=True
            )
            cell.border = border

        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(
                    vertical="top",
                    wrap_text=True
                )
                cell.border = border

        for column_cells in sheet.columns:
            max_length = max(
                len(str(cell.value or ""))
                for cell in column_cells
            )

            sheet.column_dimensions[
                column_cells[0].column_letter
            ].width = min(
                max_length + 2,
                60
            )


def export_testcases_to_excel(
    ticket_id: str,
    analysis: dict,
    scenarios: list,
    testcases: list,
    coverage_review: dict,
    final_review: dict,
    clarifications: dict | None = None,
    clarification_answers: dict | None = None,
    version: str = "latest",
    improvement_history: list | None = None,
):

    output_dir = (
        Path("requirements")
        / ticket_id
        / "exports"
    )

    output_dir.mkdir(
        parents=True,
        exist_ok=True
    )

    output_file = (
        output_dir
        / f"testcases_{version}.xlsx"
    )

    wb = Workbook()

    requirement_lookup = _build_requirement_lookup(
        analysis
    )

    # Sheet 1: Requirements
    ws_req = wb.active
    ws_req.title = "Requirements"

    ws_req.append(
        [
            "Requirement ID",
            "Type",
            "Description"
        ]
    )

    for item in analysis.get(
        "requirement_items",
        []
    ):
        ws_req.append(
            [
                item.get("requirement_id", ""),
                item.get("type", ""),
                item.get("description", "")
            ]
        )
    
    # Sheet 2: Clarifications 
    ws_clarifications = wb.create_sheet(
        "Clarifications"
    )

    ws_clarifications.append(
        [
            "Question ID",
            "Question",
            "Answer",
            "Answer Status",
            "Answered At",
            "Impact / Reason",
            "Priority",
            "Related Requirement",
            "Category",
        ]
    )

    clarifications = clarifications or {}
    clarification_answers = clarification_answers or {}

    questions = clarifications.get(
        "clarification_questions",
        []
    )

    answered = normalize_clarification_answers(clarification_answers)
    answers_by_id = {}
    answers_by_question = {}

    for item in answered:
        question_id = get_clarification_id(item)
        question_text = normalize_question_text(get_clarification_question(item))

        if question_id:
            answers_by_id[question_id] = item
        if question_text:
            answers_by_question[question_text] = item

    matched_answer_keys = set()
    matched_answer_indexes = set()

    for question in questions:

        question_id = get_clarification_id(question)
        question_text = get_clarification_question(question)
        answer_info = answers_by_id.get(question_id, {})

        if not answer_info:
            answer_info = answers_by_question.get(
                normalize_question_text(question_text),
                {},
            )

        answer = get_clarification_answer_text(answer_info) if answer_info else ""

        if answer_info:
            matched_answer_indexes.add(id(answer_info))
            answer_id = get_clarification_id(answer_info)
            answer_question = normalize_question_text(
                get_clarification_question(answer_info)
            )
            if answer_id:
                matched_answer_keys.add(("id", answer_id))
            if answer_question:
                matched_answer_keys.add(("question", answer_question))

        ws_clarifications.append(
            [
                question_id,
                question_text,
                answer,
                (
                    "Answered"
                    if answer
                    else "Unanswered"
                ),
                answer_info.get("answered_at", "") if answer_info else "",
                question.get("impact") or question.get("reason", ""),
                question.get("priority", ""),
                _to_text(
                    question.get("related_requirement")
                    or question.get("related_requirement_id")
                    or question.get("related_requirement_ids")
                    or question.get("requirement_id")
                    or ""
                ),
                question.get("category", ""),
            ]
        )

    for item in answered:
        question_id = get_clarification_id(item)
        question_text = get_clarification_question(item)
        normalized_question = normalize_question_text(question_text)

        if (
            question_id
            and ("id", question_id) in matched_answer_keys
        ) or (
            normalized_question
            and ("question", normalized_question) in matched_answer_keys
        ):
            continue

        ws_clarifications.append(
            [
                question_id,
                question_text,
                get_clarification_answer_text(item),
                "Answered",
                item.get("answered_at", ""),
                item.get("impact") or item.get("reason", ""),
                item.get("priority", ""),
                _to_text(
                    item.get("related_requirement")
                    or item.get("related_requirement_id")
                    or item.get("related_requirement_ids")
                    or item.get("requirement_id")
                    or ""
                ),
                item.get("category", ""),
            ]
        )

    clarification_export_stats = {
        "clarification_count": len(questions),
        "answer_count": len(answered),
        "matched_answer_count": len(matched_answer_indexes),
    }
        
    # Sheet 3: Requirement Coverage Matrix
    ws_matrix = wb.create_sheet(
        "Requirement Coverage Matrix"
    )

    ws_matrix.append(
        [
            "Requirement ID",
            "Type",
            "Description",
            "Covered",
            "Scenario Count",
            "Test Case Count",
            "Scenario IDs",
            "Test Case IDs"
        ]
    )

    coverage_matrix = _build_requirement_coverage_matrix(
        analysis,
        scenarios,
        testcases
    )

    for row in coverage_matrix:
        ws_matrix.append(
            [
                row.get(
                    "requirement_id",
                    ""
                ),
                row.get(
                    "type",
                    ""
                ),
                row.get(
                    "description",
                    ""
                ),
                row.get(
                    "covered",
                    ""
                ),
                row.get(
                    "scenario_count",
                    0
                ),
                row.get(
                    "testcase_count",
                    0
                ),
                _to_text(
                    row.get(
                        "scenario_ids",
                        []
                    )
                ),
                _to_text(
                    row.get(
                        "testcase_ids",
                        []
                    )
                )
            ]
        )

    # Sheet 4: Scenarios
    ws_scenario = wb.create_sheet(
        "Scenarios"
    )

    ws_scenario.append(
        [
            "Scenario ID",
            "Title",
            "Category",
            "Description",
            "Related Requirement IDs",
            "Related Requirement Details"
        ]
    )

    for scenario in scenarios:
        related_ids = scenario.get(
            "related_requirement_ids",
            []
        )

        ws_scenario.append(
            [
                scenario.get("scenario_id", ""),
                scenario.get("title", ""),
                scenario.get("category", ""),
                scenario.get("description", ""),
                _to_text(related_ids),
                _format_requirement_details(
                    related_ids,
                    requirement_lookup
                )
            ]
        )

    # Sheet 5: Test Cases
    ws_tc = wb.create_sheet(
        "Test Cases"
    )

    ws_tc.append(
        [
            "Test Case ID",
            "Scenario ID",
            "Traceability",
            "Related Requirement Details",
            "Title",
            "Priority",
            "Preconditions",
            "Test Steps",
            "Expected Results"
        ]
    )

    for testcase in testcases:
        related_ids = testcase.get(
            "related_requirement_ids",
            []
        )
        
        #print(
        #    "DEBUG related_ids =",
        #    related_ids
        #)
        
        #traceability = ", ".join(
        #    related_ids
        #)
        
        traceability = ", ".join(
            item["id"]
            if isinstance(item, dict)
            else str(item)
            for item in related_ids
        )

        ws_tc.append(
            [
                testcase.get("testcase_id", ""),
                testcase.get("scenario_id", ""),
                traceability,
                _format_requirement_details(
                    related_ids,
                    requirement_lookup
                ),
                testcase.get("title", ""),
                testcase.get("priority", ""),
                _to_text(
                    testcase.get(
                        "preconditions",
                        []
                    )
                ),
                _to_text(
                    testcase.get(
                        "test_steps",
                        []
                    )
                ),
                _to_text(
                    testcase.get(
                        "expected_results",
                        []
                    )
                )
            ]
        )

    # Sheet 6: Coverage Review
    ws_review = wb.create_sheet(
        "Coverage Review"
    )

    ws_review.append(
        [
            "Metric",
            "Value"
        ]
    )

    ws_review.append(
        [
            "Coverage Score",
            coverage_review.get(
                "coverage_score",
                ""
            )
        ]
    )

    ws_review.append(
        [
            "Missing Coverage",
            _to_text(
                coverage_review.get(
                    "missing_coverage",
                    []
                )
            )
        ]
    )

    ws_review.append(
        [
            "Recommendations",
            _to_text(
                coverage_review.get(
                    "recommendations",
                    []
                )
            )
        ]
    )

    # Sheet 7: Final Review
    ws_final = wb.create_sheet(
        "Final Review"
    )

    ws_final.append(
        [
            "Metric",
            "Value"
        ]
    )

    ws_final.append(
        [
            "Coverage Score",
            final_review.get(
                "coverage_score",
                ""
            )
        ]
    )

    ws_final.append(
        [
            "Improvement Score",
            final_review.get(
                "improvement_score",
                ""
            )
        ]
    )

    ws_final.append(
        [
            "Coverage Summary",
            final_review.get(
                "coverage_summary",
                ""
            )
        ]
    )

    ws_final.append(
        [
            "Remaining Gaps",
            _to_text(
                final_review.get(
                    "remaining_gaps",
                    []
                )
            )
        ]
    )

    ws_final.append(
        [
            "Recommendations",
            _to_text(
                final_review.get(
                    "recommendations",
                    []
                )
            )
        ]
    )
    
    # Sheet 8: Improvement History
    ws_history = wb.create_sheet(
        "Improvement History"
    )

    ws_history.append(
        [
            "Version",
            "Iteration",
            "Coverage Score",
            "Improvement Score",
            "Note"
        ]
    )

    improvement_history = improvement_history or []

    for item in improvement_history:
        ws_history.append(
            [
                item.get("version", ""),
                item.get("iteration", ""),
                item.get("coverage_score", ""),
                item.get("improvement_score", ""),
                item.get("note", "")
            ]
        )

    _apply_styles(wb)

    wb.save(output_file)
    logger.debug(
        "Exported Clarifications sheet",
        extra={
            "ticket_id": ticket_id,
            **clarification_export_stats,
            "export_path": str(output_file),
        },
    )

    return str(output_file)
