from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def _apply_styles(wb):

    header_fill = PatternFill(
        "solid",
        fgColor="D9EAF7"
    )

    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    for sheet in wb.worksheets:

        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(
                vertical="center",
                wrap_text=True
            )
            cell.border = border

        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(
                    vertical="top",
                    wrap_text=True
                )
                cell.border = border

        for column_cells in sheet.columns:
            max_length = max(
                len(str(cell.value or ""))
                for cell in column_cells
            )

            sheet.column_dimensions[
                column_cells[0].column_letter
            ].width = min(
                max_length + 2,
                70
            )


def export_test_case_structure_to_excel(
    ticket_id: str,
    structure: dict,
    review: dict
):

    output_dir = (
        Path("requirements")
        / ticket_id
        / "exports"
    )

    output_dir.mkdir(
        parents=True,
        exist_ok=True
    )

    output_file = (
        output_dir
        / "test_case_structure.xlsx"
    )

    wb = Workbook()

    ws_main = wb.active
    ws_main.title = "Main Functions"

    ws_main.append(
        [
            "Function ID",
            "Function Name",
            "Description",
            "Related Requirement IDs"
        ]
    )

    ws_sub = wb.create_sheet(
        "Sub Functions"
    )

    ws_sub.append(
        [
            "Function ID",
            "Function Name",
            "Sub Function ID",
            "Sub Function Name",
            "Description",
            "Related Requirement IDs"
        ]
    )

    ws_cat = wb.create_sheet(
        "Test Categories"
    )

    ws_cat.append(
        [
            "Function ID",
            "Function Name",
            "Sub Function ID",
            "Sub Function Name",
            "Category ID",
            "Category Name",
            "Test Intent",
            "Priority",
            "Related Requirement IDs"
        ]
    )

    for function in structure.get(
        "main_functions",
        []
    ):

        function_id = function.get(
            "function_id",
            ""
        )

        function_name = function.get(
            "function_name",
            ""
        )

        ws_main.append(
            [
                function_id,
                function_name,
                function.get("description", ""),
                ", ".join(
                    function.get(
                        "related_requirement_ids",
                        []
                    )
                )
            ]
        )

        for sub in function.get(
            "sub_functions",
            []
        ):

            sub_id = sub.get(
                "sub_function_id",
                ""
            )

            sub_name = sub.get(
                "sub_function_name",
                ""
            )

            ws_sub.append(
                [
                    function_id,
                    function_name,
                    sub_id,
                    sub_name,
                    sub.get("description", ""),
                    ", ".join(
                        sub.get(
                            "related_requirement_ids",
                            []
                        )
                    )
                ]
            )

            for category in sub.get(
                "test_categories",
                []
            ):

                ws_cat.append(
                    [
                        function_id,
                        function_name,
                        sub_id,
                        sub_name,
                        category.get(
                            "category_id",
                            ""
                        ),
                        category.get(
                            "category_name",
                            ""
                        ),
                        category.get(
                            "test_intent",
                            ""
                        ),
                        category.get(
                            "priority",
                            ""
                        ),
                        ", ".join(
                            category.get(
                                "related_requirement_ids",
                                []
                            )
                        )
                    ]
                )

    ws_review = wb.create_sheet(
        "AI Review"
    )

    ws_review.append(
        [
            "Section",
            "Content"
        ]
    )

    for key, value in review.items():

        if isinstance(value, list):
            value = "\n".join(
                str(item)
                for item in value
            )

        ws_review.append(
            [
                key,
                value
            ]
        )

    _apply_styles(wb)

    wb.save(
        output_file
    )

    return str(output_file)