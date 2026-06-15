import logging
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from app.utils.clarification_answers import (
    get_clarification_answer_text,
    get_clarification_id,
    get_clarification_question,
    normalize_clarification_answers,
    normalize_question_text,
)


logger = logging.getLogger(__name__)


def _apply_styles(wb: Workbook):

    header_fill = PatternFill(
        "solid",
        fgColor="D9EAF7"
    )

    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    for sheet in wb.worksheets:

        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(
                vertical="center",
                wrap_text=True
            )
            cell.border = border

        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(
                    vertical="top",
                    wrap_text=True
                )
                cell.border = border

        for column_cells in sheet.columns:
            max_length = max(
                len(str(cell.value or ""))
                for cell in column_cells
            )

            sheet.column_dimensions[
                column_cells[0].column_letter
            ].width = min(
                max_length + 2,
                70
            )


def export_requirement_intelligence_to_excel(
    ticket_id: str,
    analysis: dict,
    clarifications: dict,
    clarification_answers: dict,
    requirement_summary: dict
):

    output_dir = (
        Path("requirements")
        / ticket_id
        / "exports"
    )

    output_dir.mkdir(
        parents=True,
        exist_ok=True
    )

    output_file = (
        output_dir
        / "requirement_intelligence.xlsx"
    )

    wb = Workbook()

    # Sheet 1: Requirements
    ws_req = wb.active
    ws_req.title = "Requirements"

    ws_req.append(
        [
            "Requirement ID",
            "Type",
            "Description"
        ]
    )

    for item in analysis.get(
        "requirement_items",
        []
    ):
        ws_req.append(
            [
                item.get("requirement_id", ""),
                item.get("type", ""),
                item.get("description", "")
            ]
        )

    # Sheet 2: Clarifications
    ws_clar = wb.create_sheet(
        "Clarifications"
    )

    ws_clar.append(
        [
            "Question ID",
            "Question",
            "Answer",
            "Answer Status",
            "Answered At",
            "Impact / Reason",
            "Priority",
            "Related Requirement",
            "Category",
        ]
    )

    questions = clarifications.get(
        "clarification_questions",
        []
    )

    answered_clarifications = normalize_clarification_answers(
        clarification_answers
    )

    answers_by_id = {}
    answers_by_question = {}

    for item in answered_clarifications:
        question_id = get_clarification_id(item)
        question_text = normalize_question_text(get_clarification_question(item))

        if question_id:
            answers_by_id[question_id] = item
        if question_text:
            answers_by_question[question_text] = item

    matched_answer_keys = set()
    matched_answer_indexes = set()

    for question in questions:

        question_id = get_clarification_id(question)
        question_text = get_clarification_question(question)
        answer_info = answers_by_id.get(question_id, {})

        if not answer_info:
            answer_info = answers_by_question.get(
                normalize_question_text(question_text),
                {},
            )

        answer = get_clarification_answer_text(answer_info) if answer_info else ""

        status = (
            "Answered"
            if answer
            else "Unanswered"
        )

        if answer_info:
            matched_answer_indexes.add(id(answer_info))
            answer_id = get_clarification_id(answer_info)
            answer_question = normalize_question_text(
                get_clarification_question(answer_info)
            )
            if answer_id:
                matched_answer_keys.add(("id", answer_id))
            if answer_question:
                matched_answer_keys.add(("question", answer_question))

        ws_clar.append(
            [
                question_id,
                question_text,
                answer,
                status,
                answer_info.get("answered_at", "") if answer_info else "",
                question.get("impact") or question.get("reason", ""),
                question.get("priority", ""),
                question.get("related_requirement")
                or question.get("related_requirement_id")
                or question.get("related_requirement_ids")
                or question.get("requirement_id")
                or "",
                question.get("category", ""),
            ]
        )

    for item in answered_clarifications:
        question_id = get_clarification_id(item)
        question_text = get_clarification_question(item)
        normalized_question = normalize_question_text(question_text)

        if (
            question_id
            and ("id", question_id) in matched_answer_keys
        ) or (
            normalized_question
            and ("question", normalized_question) in matched_answer_keys
        ):
            continue

        ws_clar.append(
            [
                question_id,
                question_text,
                get_clarification_answer_text(item),
                "Answered",
                item.get("answered_at", ""),
                item.get("impact") or item.get("reason", ""),
                item.get("priority", ""),
                item.get("related_requirement")
                or item.get("related_requirement_id")
                or item.get("related_requirement_ids")
                or item.get("requirement_id")
                or "",
                item.get("category", ""),
            ]
        )

    clarification_export_stats = {
        "clarification_count": len(questions),
        "answer_count": len(answered_clarifications),
        "matched_answer_count": len(matched_answer_indexes),
    }

    # Sheet 3: Requirement Summary
    ws_summary = wb.create_sheet(
        "Requirement Summary"
    )

    ws_summary.append(
        [
            "Section",
            "Content"
        ]
    )

    ws_summary.append(
        [
            "Executive Summary",
            requirement_summary.get(
                "executive_summary",
                ""
            )
        ]
    )

    ws_summary.append(
        [
            "Functional Summary",
            requirement_summary.get(
                "functional_summary",
                ""
            )
        ]
    )

    for key in [
        "confirmed_business_rules",
        "validation_rules",
        "open_questions",
        "assumptions",
        "risks"
    ]:
        value = requirement_summary.get(
            key,
            []
        )

        if isinstance(value, list):
            value = "\n".join(
                str(item)
                for item in value
            )

        ws_summary.append(
            [
                key,
                value
            ]
        )

    _apply_styles(wb)

    wb.save(
        output_file
    )
    logger.debug(
        "Exported Clarifications sheet",
        extra={
            "ticket_id": ticket_id,
            **clarification_export_stats,
            "export_path": str(output_file),
        },
    )

    return str(
        output_file
    )
