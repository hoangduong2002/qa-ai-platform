import os
from typing import Any
from typing_extensions import TypedDict
from langchain_deepseek import ChatDeepSeek
from langgraph.graph import StateGraph, START, END
from app.config.env_loader import load_project_env


load_project_env()

# Khởi tạo mô hình DeepSeek nội bộ cho Agent
llm = ChatDeepSeek(
    model="deepseek-chat",
    temperature=0.1,
    timeout=120
)

class AnalystState(TypedDict):
    ticket_id: str
    raw_requirement: str
    qa_ambiguities: str
    summarized_req: str

# 🧠 NODE 1: Phân tích điểm mập mờ nghiệp vụ
def analyze_and_create_qa_node(state: AnalystState) -> dict[str, Any]:
    print("\n[Analyst Agent] 🧠 Analyzing requirements to detect ambiguities and conflicts...")
    prompt = f"""
    Role: Expert Business Analyst (BA) and Senior QA Lead.
    Task: Review the consolidated Jira ticket details provided below:
    ---
    {state['raw_requirement']}
    ---
    Identify ambiguities, gaps, or conflicts. Generate a clarification Q&A log table in Markdown format with columns:
    - **No.** | **Source Location** | **Ambiguity / Conflict Identified** | **Potential Impact / Risk** | **Recommended Q&A Clarification** | **Proposed Solution by QA**
    Everything MUST BE IN PROFESSIONAL ENGLISH. Return ONLY the Markdown table.
    """
    response = llm.invoke(prompt)
    return {"qa_ambiguities": response.content}

# 📝 NODE 2: Tổng hợp và Chuẩn hóa lại bộ Requirements
def summarize_requirements_node(state: AnalystState) -> dict[str, Any]:
    print("\n[Analyst Agent] 📝 Synthesizing and creating clean standardized Requirements text...")
    prompt = f"""
    Role: Professional Requirements Engineer and Technical Writer.
    Task: Based on the raw Jira data and the generated Q&A logs below, create a consolidated "Summarized Requirements" document.
    
    RAW DATA: {state['raw_requirement']}
    Q&A LOGS: {state['qa_ambiguities']}
    
    Structure the output in Technical English using headers:
    # 📝 CONSOLIDATED REQUIREMENTS SUMMARY ({state['ticket_id']})
    ## 1. Business Goals & Objectives
    ## 2. Comprehensive Functional Requirements
    ## 3. Technical, Security & Non-Functional Constraints
    ## 4. Critical Pending Points for Confirmation
    Return ONLY the Markdown document.
    """
    response = llm.invoke(prompt)
    return {"summarized_req": response.content}

# ==========================================
# 📊 FILE EXPORT MANAGEMENT HANDLER
# ==========================================
def export_separated_reports(ticket_key: str, qa_markdown: str, summary_markdown: str):
    excel_filename = f"{ticket_key}_QA_Clarifications.xlsx"
    md_filename = f"{ticket_key}_Summarized_Requirements.md"
    
    print("\n[File Processing] 💾 Exporting data to separate specialized files...")
    
    # --- 1. EXPORT MARKDOWN FILE (.md) ---
    try:
        with open(md_filename, "w", encoding="utf-8") as f:
            f.write(summary_markdown)
        print(f"🎉 Standardized Requirements Summary saved successfully: {os.path.abspath(md_filename)}")
    except Exception as e:
        print(f"❌ Error creating Markdown file: {e}")

    # --- 2. EXPORT INTERACTIVE EXCEL TRACKING FILE (.xlsx) ---
    try:
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        def clean_text(text: str) -> str:
            if not text: return ""
            text = re.sub(r'<br\s*/?>', '\n', text, flags=re.IGNORECASE)
            text = re.sub(r'\*\*(.*?)\*\*', r'\1', text)
            text = re.sub(r'\*(.*?)\*', r'\1', text)
            return text.strip()

        qa_lines = [line.strip() for line in qa_markdown.strip().split("\n") if line.strip()]
        qa_rows = []
        for line in qa_lines:
            if line.startswith("|") and not re.match(r'^\|[\s|:\-]*\|$', line):
                columns = [clean_text(col) for col in line.split("|")[1:-1]]
                qa_rows.append(columns)
        
        if len(qa_rows) >= 2:
            headers = qa_rows
            data_rows = qa_rows[1:]
            headers.extend(["Assignee / Confirmed Resolution", "Status"])
            
            final_data = []
            for row in data_rows:
                row.extend(["", "Open"])
                final_data.append(row)
                
            df_qa = pd.DataFrame(final_data, columns=headers)
        else:
            df_qa = pd.DataFrame(columns=["No.", "Content", "Assignee / Confirmed Resolution", "Status"])

        with pd.ExcelWriter(excel_filename, engine='openpyxl') as writer:
            df_qa.to_excel(writer, sheet_name="Q&A Tracking", index=False)
            worksheet = writer.sheets["Q&A Tracking"]
            
            header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            
            input_header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            input_header_font = Font(name="Arial", size=11, bold=True, color="1F4E78")
            
            data_font = Font(name="Arial", size=10, bold=False)
            thin_border = Border(
                left=Side(style='thin', color='E0E0E0'), right=Side(style='thin', color='E0E0E0'),
                top=Side(style='thin', color='E0E0E0'), bottom=Side(style='thin', color='E0E0E0')
            )
            
            for col_num in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=1, column=col_num)
                if col_num >= worksheet.max_column - 1:
                    cell.font = input_header_font
                    cell.fill = input_header_fill
                else:
                    cell.font = header_font
                    cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = thin_border
            
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
                for col_idx, cell in enumerate(row, start=1):
                    cell.font = data_font
                    cell.border = thin_border
                    if col_idx == worksheet.max_column:
                        cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
                    else:
                        cell.alignment = Alignment(vertical="top", wrap_text=True)
            
            for col_idx, col in enumerate(worksheet.columns, start=1):
                max_len = 0
                for cell in col:
                    lines_in_cell = str(cell.value or '').split('\n')
                    for l in lines_in_cell:
                        if len(l) > max_len: max_len = len(l)
                col_letter = get_column_letter(col_idx)
                default_width = 40 if col_idx == worksheet.max_column - 1 else 25
                worksheet.column_dimensions[col_letter].width = min(max(max_len + 4, default_width), 50)
                
        print(f"🎉 Interactive Excel Q&A Tracker generated successfully: {os.path.abspath(excel_filename)}")
    except Exception as e:
        print(f"❌ Error constructing Excel worksheet file: {e}")

# Cấu hình đồ thị LangGraph chuyên biệt cho Analyst Agent
workflow = StateGraph(AnalystState)
workflow.add_node("qa_analyzer", analyze_and_create_qa_node)
workflow.add_node("req_summarizer", summarize_requirements_node)

workflow.add_edge(START, "qa_analyzer")
workflow.add_edge("qa_analyzer", "req_summarizer")
workflow.add_edge("req_summarizer", END)

requirement_analyze_agent = workflow.compile()
