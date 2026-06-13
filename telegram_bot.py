"""
Legacy Telegram entrypoint.

Use `python -m bot.telegram_bot` for the maintained bot flow.
"""

import os
import re
import asyncio
import json
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import Application, CommandHandler, MessageHandler, CallbackQueryHandler, filters, ContextTypes

from app.config.env_loader import load_project_env

# Nạp các thành phần từ các file Agent chuyên biệt của bạn
_llm_instance = None


def get_llm():
    global _llm_instance

    if _llm_instance is None:
        from langchain_deepseek import ChatDeepSeek

        _llm_instance = ChatDeepSeek(
            model="deepseek-chat",
            temperature=0.1,
            timeout=120,
        )

    return _llm_instance


class LLMProxy:
    def invoke(self, prompt: str):
        return get_llm().invoke(prompt)


llm = LLMProxy()


def get_requirement_analyze_agent():
    from requirement_analyze_agent import requirement_analyze_agent

    return requirement_analyze_agent


def get_jira_agent_app():
    from jira_agent import jira_agent_app

    return jira_agent_app


class LegacyTestcaseAgent:
    """Small compatibility agent for the legacy root bot."""

    def invoke(self, inputs: dict) -> dict:
        requirement = (inputs or {}).get("requirement", "").strip()

        prompt = f"""
        Role: Senior QA Engineer.
        Task: Generate a professional test suite from the requirement below.

        REQUIREMENT:
        ---
        {requirement}
        ---

        Return a Markdown table with columns:
        Test Case ID | Title | Preconditions | Steps | Expected Result | Priority
        """

        response = llm.invoke(prompt)

        return {
            "final_test_cases": response.content
        }


four_steps_agent = LegacyTestcaseAgent()


def _markdown_table_rows(markdown_text: str) -> list[list[str]]:
    rows = []

    for raw_line in (markdown_text or "").splitlines():
        line = raw_line.strip()

        if not line.startswith("|"):
            continue

        if re.match(r"^\|[\s|:\-]*\|$", line):
            continue

        columns = [
            column.strip().strip("*")
            for column in line.split("|")[1:-1]
        ]

        if columns:
            rows.append(columns)

    return rows


def export_test_excel(final_test_cases: str, output_filename: str):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Test Cases"

    rows = _markdown_table_rows(final_test_cases)

    if rows:
        for row in rows:
            worksheet.append(row)
    else:
        worksheet.append(["Generated Test Cases"])
        for line in (final_test_cases or "").splitlines():
            if line.strip():
                worksheet.append([line.strip()])

    for cell in worksheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(
            start_color="1F4E78",
            end_color="1F4E78",
            fill_type="solid",
        )

    for row in worksheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )

    for column_cells in worksheet.columns:
        max_length = max(
            len(str(cell.value or ""))
            for cell in column_cells
        )
        worksheet.column_dimensions[column_cells[0].column_letter].width = min(
            max(max_length + 4, 18),
            60,
        )

    workbook.save(output_filename)


def export_jira_reports(ticket_id: str, qa_markdown: str, summary_markdown: str):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    md_file = f"{ticket_id}_Summarized_Requirements.md"
    excel_file = f"{ticket_id}_QA_Clarifications.xlsx"

    with open(md_file, "w", encoding="utf-8") as file:
        file.write(summary_markdown or "")

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Q&A Tracking"

    rows = _markdown_table_rows(qa_markdown)

    if rows:
        for row in rows:
            worksheet.append(row)
    else:
        worksheet.append([
            "No.",
            "Content",
            "Assignee / Confirmed Resolution",
            "Status",
        ])
        worksheet.append(["1", qa_markdown or "", "", "Open"])

    first_empty_column = worksheet.max_column + 1
    worksheet.cell(
        row=1,
        column=first_empty_column,
        value="Assignee / Confirmed Resolution",
    )
    worksheet.cell(
        row=1,
        column=first_empty_column + 1,
        value="Status",
    )

    for row_index in range(2, worksheet.max_row + 1):
        worksheet.cell(
            row=row_index,
            column=first_empty_column + 1,
            value="Open",
        )

    for cell in worksheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(
            start_color="1F4E78",
            end_color="1F4E78",
            fill_type="solid",
        )

    for row in worksheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )

    for column_cells in worksheet.columns:
        max_length = max(
            len(str(cell.value or ""))
            for cell in column_cells
        )
        worksheet.column_dimensions[column_cells[0].column_letter].width = min(
            max(max_length + 4, 18),
            60,
        )

    workbook.save(excel_file)

# Nạp cấu hình môi trường từ file .env
load_project_env()

# ==========================================
# 🤖 HÀM HIỂN THỊ MENU HƯỚNG DẪN BAN ĐẦU
# ==========================================
async def start_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    welcome_text = (
        "🤖 **Welcome to Smart Multi-Agent Assistant!**\n\n"
        "Tôi sử dụng công nghệ Router Agent thông minh của DeepSeek. "
        "Bạn không cần gõ lệnh phức tạp, hãy cứ chat tự nhiên với tôi như đồng nghiệp:\n\n"
        "💡 **Ví dụ câu lệnh bạn có thể chat:**\n"
        "1. *'Hãy phân tích giúp tôi ticket SEC-102 và lên danh sách Q&A nhé'* \n"
        "2. *'Viết test case chi tiết cho yêu cầu sau: Hệ thống cần tính năng quên mật khẩu...'* \n"
        "3. *'Tạo bộ test cases dựa trên dữ liệu của task Jira PROJ-456'*\n\n"
        "📌 *Hệ thống sẽ hiển thị nút bấm để bạn chọn có muốn đọc các sub-tickets hay không.*"
    )
    await update.message.reply_text(welcome_text, parse_mode="Markdown")


# ==========================================
# 🧠 SMART ROUTER: ĐỌC VỊ Ý ĐỊNH VÀ TRÍCH XUẤT THÔNG TIN
# ==========================================
async def smart_router_agent(user_message: str) -> dict:
    prompt = f"""
    You are an AI Router Agent for a software development chat channel. Your job is to analyze the user's request and determine which specialist agent should handle it, along with extracting necessary parameters.

    USER MESSAGE: "{user_message}"

    Analyze the message and classify it into one of these 2 core intents:
    1. "JIRA_INTENT": The user wants to perform an action on a Jira ticket (either analyzing it for Q&A/Summary OR generating test cases from it).
    2. "GENERATE_TESTCASE_FROM_TEXT": The user pasted a raw text requirement directly and wants to generate automated test cases from it (No Jira ticket involved).

    Also, extract:
    - "ticket_id": If a Jira ticket key (like SEC-102, PROJ-45) is mentioned.
    - "raw_requirement": If they pasted a text-based requirement directly into the message.
    - "sub_intent": If it's a JIRA_INTENT, specify if they want "ANALYSIS" (Q&A/Summary) or "TESTCASE" (Write test suite).

    Return ONLY a clean JSON object without markdown block formatting. Example format:
    {{"route": "JIRA_INTENT", "ticket_id": "SEC-102", "sub_intent": "ANALYSIS", "raw_requirement": ""}}
    """
    
    loop = asyncio.get_event_loop()
    response = await loop.run_in_executor(None, llm.invoke, prompt)
    clean_json_str = response.content.strip().replace("```json", "").replace("```", "")
    
    try:
        return json.loads(clean_json_str)
    except Exception:
        if "SEC-" in user_message.upper() or "PROJ-" in user_message.upper() or "EVNWCL-" in user_message.upper() or "JIRA" in user_message.upper():
            match = re.search(r'([A-Z0-9]+-\d+)', user_message.upper())
            ticket = match.group(1) if match else ""
            sub = "TESTCASE" if "TEST" in user_message.upper() else "ANALYSIS"
            return {"route": "JIRA_INTENT", "ticket_id": ticket, "sub_intent": sub, "raw_requirement": ""}
        return {"route": "GENERATE_TESTCASE_FROM_TEXT", "ticket_id": "", "sub_intent": "", "raw_requirement": user_message}


# ==========================================
# ⚡ PHÒNG ĐIỀU PHỐI TIN NHẮN TRUNG TÂM & TẠO NÚT BẤM
# ==========================================
async def message_central_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not update.message or not update.message.text:
        return 

    user_text = update.message.text.strip()
    loop = asyncio.get_event_loop()
    
    status_msg = await update.message.reply_text("🧠 *AI Assistant is thinking and routing your request...*", parse_mode="Markdown")
    routing_decision = await smart_router_agent(user_text)
    
    route = routing_decision.get("route")
    ticket_id = routing_decision.get("ticket_id", "").upper()
    sub_intent = routing_decision.get("sub_intent", "ANALYSIS")
    extracted_requirement = routing_decision.get("raw_requirement", "")
    
    # 📌 NẾU LIÊN QUAN ĐẾN JIRA -> HIỂN THỊ NÚT BẤM KHẢO SÁT SUB-TICKET
    if route == "JIRA_INTENT":
        if not ticket_id:
            await status_msg.edit_text("⚠️ Router detected JIRA intent, but could not find a valid Ticket ID in your message.")
            return
            
        # Lưu thông tin vào user_data để tái sử dụng ở hàm xử lý bấm nút riêng biệt
        context.user_data["current_ticket_id"] = ticket_id
        context.user_data["current_sub_intent"] = sub_intent
        
        # SỬA ĐOẠN CODE TRONG HÌNH ẢNH CỦA BẠN TẠI ĐÂY
        keyboard = [
            [
                InlineKeyboardButton("✅ Yes, include sub-tickets", callback_data="subtask_yes"),
                InlineKeyboardButton("❌ No, main ticket only", callback_data="subtask_no")
            ]
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)
        
        await status_msg.delete() 
        await update.message.reply_text(
            f"🌿 Task initialized for Ticket **{ticket_id}** ({sub_intent} Mode).\n"
            f"Do you want the Agent to scrape and process all attached **Sub-tickets (Sub-tasks)**?",
            reply_markup=reply_markup,
            parse_mode="Markdown"
        )

    # ✍️ NẾU LÀ VIẾT TEST CASE TỪ TEXT PHẲNG -> CHẠY LUÔN KHÔNG CẦN HỎI
    elif route == "GENERATE_TESTCASE_FROM_TEXT":
        req_data = extracted_requirement if extracted_requirement else user_text
        await status_msg.edit_text("🧠 *[Test Case Agent]* Writing test cases from your text using 4-Step prompting matrix...")
        
        initial_inputs = {"requirement": req_data}
        final_outputs = await loop.run_in_executor(None, four_steps_agent.invoke, initial_inputs)
        
        output_filename = "Text_Based_Test_Suite.xlsx"
        await loop.run_in_executor(None, export_test_excel, final_outputs["final_test_cases"], output_filename)
        
        await status_msg.edit_text("🚀 Sending test cases...")
        if os.path.exists(output_filename):
            with open(output_filename, 'rb') as doc: 
                await update.message.reply_document(document=doc, filename=output_filename, caption="📊 Test Suite from text specification")
        await status_msg.delete()
        
    else:
        await status_msg.edit_text("🤖 Sorry, I could not confidently determine your request's intent. Please try phrasing it differently or type /help.")
# ==========================================
# 🖱️ HÀM XỬ LÝ SỰ KIỆN CLICK NÚT TRÊN TELEGRAM (CALLBACK QUERY)
# ==========================================
async def handle_subtask_selection(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer() 
    
    ticket_id = context.user_data.get("current_ticket_id")
    sub_intent = context.user_data.get("current_sub_intent", "ANALYSIS")
    user_choice = query.data 
    
    if not ticket_id:
        await query.edit_message_text("⚠️ Session expired or invalid data. Please resend the Jira Ticket key.")
        return
        
    loop = asyncio.get_event_loop()
    
    if user_choice == "subtask_yes":
        status_text = f"🚀 Starting workflow for **{ticket_id}**... Mode: **Recursive (With Sub-tickets)**."
    else:
        status_text = f"🚀 Starting workflow for **{ticket_id}**... Mode: **Main Ticket Only**."
        
    await query.edit_message_text(text=status_text, parse_mode="Markdown")
    status_msg = query.message 

    try:
        # 🔌 BƯỚC 1: GỌI AGENT 1 (JIRA FETCHER) ĐỂ CÀO DATA SẠCH THEO LỰA CHỌN
        await status_msg.edit_text(f"🔌 *[Jira Fetcher]* Gathering configurations for **{ticket_id}**...", parse_mode="Markdown")
        
        if user_choice == "subtask_yes":
            initial_inputs = {"ticket_id": ticket_id, "subtask_queue": [], "current_context": ""}
        else:
            initial_inputs = {"ticket_id": ticket_id, "subtask_queue": ["SKIP_LOOP"], "current_context": ""}
            
        jira_agent_app = get_jira_agent_app()
        jira_outputs = await loop.run_in_executor(None, jira_agent_app.invoke, initial_inputs)
        compiled_context = jira_outputs["current_context"].replace("=== SUB-TICKETS LẶP: SKIP_LOOP ===", "")

        # 🎯 BƯỚC 2: ĐIỀU PHỐI ĐẾN ĐÚNG CHUYÊN GIA (ANALYST HOẶC TESTCASE AGENT)
        
        # Tuyến A: Phân tích nghiệp vụ (Gọi sang file requirement_analyze_agent.py độc lập mới tách)
        if sub_intent == "ANALYSIS":
            await status_msg.edit_text("🧠 *[Requirement Analyst Agent]* Analyzing gaps, building Q&A and technical summary...")
            analyst_inputs = {"ticket_id": ticket_id, "raw_requirement": compiled_context}
            requirement_analyze_agent = get_requirement_analyze_agent()
            analyst_outputs = await loop.run_in_executor(None, requirement_analyze_agent.invoke, analyst_inputs)
            
            await status_msg.edit_text("💾 *[System]* Generating clean Excel tracker and Markdown summary report...")
            await loop.run_in_executor(None, export_jira_reports, ticket_id, analyst_outputs["qa_ambiguities"], analyst_outputs["summarized_req"])
            
            excel_file = f"{ticket_id}_QA_Clarifications.xlsx"
            md_file = f"{ticket_id}_Summarized_Requirements.md"
            
            await status_msg.edit_text("🚀 Sending Jira Analysis reports...")
            if os.path.exists(md_file):
                with open(md_file, 'rb') as doc: await update.effective_message.reply_document(document=doc, filename=md_file, caption=f"📝 Summary for {ticket_id}")
            if os.path.exists(excel_file):
                with open(excel_file, 'rb') as doc: await update.effective_message.reply_document(document=doc, filename=excel_file, caption=f"📊 Q&A Tracker for {ticket_id}")
            await status_msg.delete()

        # Tuyến B: Chuyển tiếp kho dữ liệu sạch sang cho Agent 3 viết Test Case 4 bước tiếng Anh
        elif sub_intent == "TESTCASE":
            await status_msg.edit_text("🧠 *[Test Case Agent]* Injecting compiled Jira documentation into 4-Step matrix...")
            
            test_inputs = {"requirement": compiled_context}
            test_outputs = await loop.run_in_executor(None, four_steps_agent.invoke, test_inputs)
            
            final_test_filename = f"{ticket_id}_Generated_Test_Suite.xlsx"
            await loop.run_in_executor(None, export_test_excel, test_outputs["final_test_cases"], final_test_filename)
            
            await status_msg.edit_text(f"🚀 Sending comprehensive test suite for {ticket_id}...")
            if os.path.exists(final_test_filename):
                with open(final_test_filename, 'rb') as doc:
                    await update.effective_message.reply_document(document=doc, filename=final_test_filename, caption=f"📊 Test Suite for {ticket_id}")
            await status_msg.delete()

    except Exception as e:
        await status_msg.edit_text(f"❌ Error occurred while executing request for {ticket_id}.\nDetails: {e}")
        
    context.user_data.clear()


# ==========================================
# CẤU HÌNH VÀ KÍCH HOẠT SERVER BOT TELEGRAM
# ==========================================
if __name__ == "__main__":
    bot_token = os.getenv("TELEGRAM_BOT_TOKEN")
    if not bot_token:
        print("❌ TELEGRAM_BOT_TOKEN missing in .env!")
    else:
        print("🚀 Smart Routing Multi-Agent Server (with Interactive Option) is running...")
        app = Application.builder().token(bot_token).build()
        
        app.add_handler(CommandHandler("start", start_command))
        app.add_handler(CommandHandler("help", start_command))
        
        # Đăng ký hàm lắng nghe sự kiện bấm nút tương tác
        app.add_handler(CallbackQueryHandler(handle_subtask_selection))
        
        # Lắng nghe tin nhắn chữ tự nhiên từ người dùng
        app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, message_central_handler))
        app.run_polling()
